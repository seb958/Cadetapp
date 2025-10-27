from fastapi import FastAPI, APIRouter, HTTPException, Depends, status, UploadFile, File
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
from pathlib import Path
import re
import secrets
from pydantic import BaseModel, Field, EmailStr
from typing import List, Optional, Dict, Any
import uuid
import unicodedata
from datetime import datetime, timedelta, date, timezone
from passlib.context import CryptContext
import jwt
from enum import Enum
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from io import BytesIO
from fastapi.responses import StreamingResponse
import pandas as pd

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

# Fonctions utilitaires pour la gestion des usernames
def normalize_text(text: str) -> str:
    """Normalise le texte en supprimant les accents et caractères spéciaux"""
    # Normaliser les caractères Unicode (décomposer les accents)
    normalized = unicodedata.normalize('NFD', text)
    # Supprimer les caractères de catégorie "Mark" (accents)
    ascii_text = ''.join(char for char in normalized if unicodedata.category(char) != 'Mn')
    # Convertir en minuscules et supprimer les caractères non alphanumériques
    return re.sub(r'[^a-z0-9]', '', ascii_text.lower())

def generate_base_username(prenom: str, nom: str) -> str:
    """Génère un username de base à partir du prénom et nom"""
    # Prendre la première lettre du prénom + nom complet
    prenom_normalized = normalize_text(prenom)
    nom_normalized = normalize_text(nom)
    
    if prenom_normalized and nom_normalized:
        return f"{prenom_normalized[0]}{nom_normalized}"
    elif nom_normalized:
        return nom_normalized
    else:
        return "user"

async def generate_unique_username(prenom: str, nom: str) -> str:
    """Génère un username unique en ajoutant un chiffre si nécessaire"""
    base_username = generate_base_username(prenom, nom)
    
    # Vérifier si le username de base existe déjà
    existing_user = await db.users.find_one({"username": base_username})
    
    if not existing_user:
        return base_username
    
    # Si le username existe, ajouter un chiffre
    counter = 2
    while True:
        new_username = f"{base_username}{counter}"
        existing_user = await db.users.find_one({"username": new_username})
        
        if not existing_user:
            return new_username
        
        counter += 1
        
        # Éviter les boucles infinies
        if counter > 100:
            # Utiliser un UUID aléatoire en dernier recours
            return f"{base_username}{secrets.randbelow(10000)}"

# Security
SECRET_KEY = os.environ.get("SECRET_KEY", "your-secret-key-change-in-production")
ALGORITHM = "HS256"
ACCESS_TOKEN_EXPIRE_MINUTES = 480  # 8 heures au lieu de 30 minutes

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
security = HTTPBearer()

# Create the main app
app = FastAPI(title="Gestion Escadron Cadets", version="1.0.0")
api_router = APIRouter(prefix="/api")

# Enums
class UserRole(str, Enum):
    CADET = "cadet"
    CADET_RESPONSIBLE = "cadet_responsible" 
    CADET_ADMIN = "cadet_admin"
    ENCADREMENT = "encadrement"

class Grade(str, Enum):
    # Grades pour cadets (ordre croissant)
    CADET = "cadet"
    CADET_AIR_1RE_CLASSE = "cadet_air_1re_classe"
    CAPORAL = "caporal"
    CAPORAL_SECTION = "caporal_section"
    SERGENT = "sergent"
    SERGENT_SECTION = "sergent_section"
    ADJUDANT_2E_CLASSE = "adjudant_2e_classe"
    ADJUDANT_1RE_CLASSE = "adjudant_1re_classe"
    INSTRUCTEUR_CIVIL = "instructeur_civil"
    ELEVE_OFFICIER = "eleve_officier"
    SOUS_LIEUTENANT = "sous_lieutenant"
    LIEUTENANT = "lieutenant"
    CAPITAINE = "capitaine"
    # Grades legacy (pour compatibilité avec données existantes)
    ADJUDANT = "adjudant"  # Legacy
    COMMANDANT = "commandant"  # Legacy

class PresenceStatus(str, Enum):
    PRESENT = "present"
    ABSENT = "absent"
    RETARD = "retard"

class ActivityType(str, Enum):
    UNIQUE = "unique"  # Activité ponctuelle
    RECURRING = "recurring"  # Activité récurrente

# Models
class UserBase(BaseModel):
    nom: str
    prenom: str
    username: Optional[str] = None  # Nom d'utilisateur pour connexion (ex: sdesy)
    email: Optional[EmailStr] = None
    grade: Grade
    role: str  # Changé de UserRole à str pour supporter les rôles personnalisés
    section_id: Optional[str] = None
    subgroup_id: Optional[str] = None  # Sous-groupe optionnel dans la section
    photo_base64: Optional[str] = None
    actif: bool = True
    has_admin_privileges: bool = False  # Privilège "cadet admin" en plus du rôle

class UserCreate(UserBase):
    pass

class UserInDB(UserBase):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    hashed_password: Optional[str] = None
    invitation_token: Optional[str] = None
    invitation_expires: Optional[datetime] = None
    created_at: datetime = Field(default_factory=datetime.utcnow)
    created_by: Optional[str] = None
    must_change_password: bool = False  # Force le changement de mot de passe à la prochaine connexion

class User(UserBase):
    id: str
    created_at: datetime
    must_change_password: bool = False

class UserInvitation(BaseModel):
    email: Optional[EmailStr] = None
    nom: str
    prenom: str
    grade: Grade
    role: str  # Changé de UserRole à str pour supporter les rôles personnalisés
    section_id: Optional[str] = None
    subgroup_id: Optional[str] = None
    has_admin_privileges: bool = False

class SetPasswordRequest(BaseModel):
    token: str
    password: str

class ChangePasswordRequest(BaseModel):
    old_password: str
    new_password: str

class GeneratePasswordResponse(BaseModel):
    user_id: str
    username: str
    temporary_password: str
    message: str

class UserUpdate(BaseModel):
    nom: Optional[str] = None
    prenom: Optional[str] = None
    email: Optional[EmailStr] = None
    grade: Optional[Grade] = None
    role: Optional[str] = None  # Changé de UserRole à str pour supporter les rôles personnalisés
    section_id: Optional[str] = None
    subgroup_id: Optional[str] = None
    actif: Optional[bool] = None
    has_admin_privileges: Optional[bool] = None

class LoginRequest(BaseModel):
    username: str  # Maintenant on utilise username au lieu d'email
    password: str

class Token(BaseModel):
    access_token: str
    token_type: str
    user: User

class Section(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    nom: str
    description: Optional[str] = None
    responsable_id: Optional[str] = None
    created_at: datetime = Field(default_factory=datetime.utcnow)

class SectionCreate(BaseModel):
    nom: str
    description: Optional[str] = None
    responsable_id: Optional[str] = None

# Modèles pour les sous-groupes
class SubGroup(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    nom: str
    description: Optional[str] = None
    section_id: str  # ID de la section parente
    responsable_id: Optional[str] = None  # Commandant de section
    created_at: datetime = Field(default_factory=datetime.utcnow)

class SubGroupCreate(BaseModel):
    nom: str
    description: Optional[str] = None
    section_id: str
    responsable_id: Optional[str] = None

class SubGroupUpdate(BaseModel):
    nom: Optional[str] = None
    description: Optional[str] = None
    responsable_id: Optional[str] = None

# Modèles pour les présences
class Presence(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    cadet_id: str
    date: date
    status: PresenceStatus
    commentaire: Optional[str] = None
    enregistre_par: str  # ID de l'utilisateur qui a enregistré
    heure_enregistrement: datetime = Field(default_factory=datetime.utcnow)
    section_id: Optional[str] = None
    activite: Optional[str] = None  # Description de l'activité
    is_guest: bool = False  # Indique si c'est un invité
    guest_nom: Optional[str] = None  # Nom de l'invité
    guest_prenom: Optional[str] = None  # Prénom de l'invité

class PresenceCreate(BaseModel):
    cadet_id: str
    status: PresenceStatus
    commentaire: Optional[str] = None
    is_guest: bool = False
    guest_nom: Optional[str] = None
    guest_prenom: Optional[str] = None

class PresenceUpdate(BaseModel):
    status: Optional[PresenceStatus] = None
    commentaire: Optional[str] = None

class PresenceBulkCreate(BaseModel):
    date: date
    activite: Optional[str] = None
    presences: List[PresenceCreate]

class PresenceResponse(BaseModel):
    id: str
    cadet_id: str
    cadet_nom: str
    cadet_prenom: str
    date: date
    status: PresenceStatus
    commentaire: Optional[str]
    enregistre_par: str
    heure_enregistrement: datetime
    section_id: Optional[str]
    section_nom: Optional[str]
    activite: Optional[str]
    is_guest: bool = False
    guest_nom: Optional[str] = None
    guest_prenom: Optional[str] = None

class PresenceStats(BaseModel):
    total_seances: int
    presences: int
    absences: int
    absences_excusees: int
    retards: int
    taux_presence: float

# Modèles pour les activités pré-définies
class Activity(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    nom: str
    description: Optional[str] = None
    type: ActivityType
    cadet_ids: List[str]  # Liste des IDs des cadets participants
    
    # Pour les activités récurrentes
    recurrence_interval: Optional[int] = None  # ex: 14 pour toutes les 2 semaines
    recurrence_unit: Optional[str] = None  # "days", "weeks", "months"
    next_date: Optional[str] = None  # Format: YYYY-MM-DD
    
    # Pour les activités ponctuelles  
    planned_date: Optional[str] = None  # Format: YYYY-MM-DD
    
    created_by: str
    created_at: datetime = Field(default_factory=datetime.utcnow)
    active: bool = True

class ActivityCreate(BaseModel):
    nom: str
    description: Optional[str] = None
    type: ActivityType
    cadet_ids: List[str]
    
    # Pour récurrence
    recurrence_interval: Optional[int] = None
    recurrence_unit: Optional[str] = None
    next_date: Optional[str] = None  # Format: YYYY-MM-DD
    
    # Pour ponctuel
    planned_date: Optional[str] = None  # Format: YYYY-MM-DD

class ActivityResponse(BaseModel):
    id: str
    nom: str
    description: Optional[str]
    type: ActivityType
    cadet_ids: List[str]
    cadet_names: List[str]  # Noms complets des cadets
    recurrence_interval: Optional[int]
    recurrence_unit: Optional[str]
    next_date: Optional[str]  # Format: YYYY-MM-DD
    planned_date: Optional[str]  # Format: YYYY-MM-DD
    created_by: str
    created_at: datetime
    active: bool

class AlertStatus(str, Enum):
    ACTIVE = "active"
    CONTACTED = "contacted"
    RESOLVED = "resolved"

class Alert(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    cadet_id: str
    consecutive_absences: int
    last_absence_date: date
    status: AlertStatus = AlertStatus.ACTIVE
    contacted_by: Optional[str] = None
    contacted_at: Optional[datetime] = None
    contact_comment: Optional[str] = None
    resolved_by: Optional[str] = None
    resolved_at: Optional[datetime] = None
    created_at: datetime = Field(default_factory=datetime.utcnow)

class AlertResponse(BaseModel):
    id: str
    cadet_id: str
    cadet_nom: str
    cadet_prenom: str
    consecutive_absences: int
    last_absence_date: date
    status: AlertStatus
    contacted_by: Optional[str] = None
    contacted_at: Optional[datetime] = None
    contact_comment: Optional[str] = None
    resolved_by: Optional[str] = None
    resolved_at: Optional[datetime] = None
    created_at: datetime

class AlertUpdate(BaseModel):
    status: AlertStatus
    contact_comment: Optional[str] = None

class Permission(str, Enum):
    # Permissions de base
    VIEW_USERS = "view_users"
    CREATE_USERS = "create_users"
    EDIT_USERS = "edit_users"
    DELETE_USERS = "delete_users"
    
    # Permissions des sections
    VIEW_SECTIONS = "view_sections"
    CREATE_SECTIONS = "create_sections"
    EDIT_SECTIONS = "edit_sections"
    DELETE_SECTIONS = "delete_sections"
    
    # Permissions des activités
    VIEW_ACTIVITIES = "view_activities"
    CREATE_ACTIVITIES = "create_activities"
    EDIT_ACTIVITIES = "edit_activities"
    DELETE_ACTIVITIES = "delete_activities"
    
    # Permissions des présences
    VIEW_PRESENCES = "view_presences"
    CREATE_PRESENCES = "create_presences"
    EDIT_PRESENCES = "edit_presences"
    DELETE_PRESENCES = "delete_presences"
    
    # Permissions des alertes
    VIEW_ALERTS = "view_alerts"
    MANAGE_ALERTS = "manage_alerts"
    
    # Permissions administratives
    MANAGE_ROLES = "manage_roles"
    SYSTEM_SETTINGS = "system_settings"

class Role(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    description: Optional[str] = None
    permissions: List[Permission] = []
    is_system_role: bool = False  # Les rôles système ne peuvent pas être supprimés
    created_at: datetime = Field(default_factory=datetime.utcnow)

class RoleCreate(BaseModel):
    name: str
    description: Optional[str] = None
    permissions: List[Permission] = []

class RoleUpdate(BaseModel):
    name: Optional[str] = None
    description: Optional[str] = None
    permissions: Optional[List[Permission]] = None

class ConsecutiveAbsenceCalculation(BaseModel):
    cadet_id: str
    consecutive_absences: int
    last_absence_date: Optional[date] = None
# Password utilities
def verify_password(plain_password: str, hashed_password: str) -> bool:
    return pwd_context.verify(plain_password, hashed_password)

def get_password_hash(password: str) -> str:
    return pwd_context.hash(password)

def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    if expires_delta:
        expire = datetime.utcnow() + expires_delta
    else:
        expire = datetime.utcnow() + timedelta(minutes=15)
    to_encode.update({"exp": expire})
    encoded_jwt = jwt.encode(to_encode, SECRET_KEY, algorithm=ALGORITHM)
    return encoded_jwt

def create_invitation_token(email: str) -> str:
    data = {
        "email": email,
        "type": "invitation",
        "exp": datetime.utcnow() + timedelta(days=7)  # Token valide 7 jours
    }
    return jwt.encode(data, SECRET_KEY, algorithm=ALGORITHM)

async def get_current_user(credentials: HTTPAuthorizationCredentials = Depends(security)):
    credentials_exception = HTTPException(
        status_code=status.HTTP_401_UNAUTHORIZED,
        detail="Token d'authentification invalide",
        headers={"WWW-Authenticate": "Bearer"},
    )
    try:
        payload = jwt.decode(credentials.credentials, SECRET_KEY, algorithms=[ALGORITHM])
        user_id: str = payload.get("sub")
        if user_id is None:
            raise credentials_exception
    except jwt.PyJWTError:
        raise credentials_exception
    
    user = await db.users.find_one({"id": user_id})
    if user is None:
        raise credentials_exception
    return User(**user)

async def require_admin_or_encadrement(current_user: User = Depends(get_current_user)):
    if current_user.role not in [UserRole.CADET_ADMIN, UserRole.ENCADREMENT]:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Accès refusé. Permissions administrateur requises."
        )
    return current_user

async def require_presence_permissions(current_user: User = Depends(get_current_user)):
    """
    Vérifie les permissions pour la gestion des présences
    Autorisé: CADET_RESPONSIBLE, CADET_ADMIN, ENCADREMENT, ou cadets avec has_admin_privileges=True
    """
    # Vérifier les rôles système
    if current_user.role in [UserRole.CADET_RESPONSIBLE, UserRole.CADET_ADMIN, UserRole.ENCADREMENT]:
        return current_user
    
    # Vérifier si l'utilisateur a le privilège admin optionnel
    if current_user.has_admin_privileges:
        return current_user
    
    raise HTTPException(
        status_code=status.HTTP_403_FORBIDDEN,
        detail="Accès refusé. Permissions pour gestion des présences requises."
    )
    return current_user

async def require_inspection_permissions(current_user: User = Depends(get_current_user)):
    """
    Vérifie les permissions pour l'inspection des uniformes
    Autorisé: Chefs de section et supérieurs
    """
    # Récupérer les rôles personnalisés qui peuvent inspecter
    allowed_system_roles = [UserRole.CADET_RESPONSIBLE, UserRole.CADET_ADMIN, UserRole.ENCADREMENT]
    
    # Rôles personnalisés autorisés (contenant "chef", "sergent", "adjudant", "officier")
    allowed_custom_keywords = ["chef", "sergent", "adjudant", "officier", "commandant"]
    
    # Vérifier rôles système
    if current_user.role in [r.value for r in allowed_system_roles]:
        return current_user
    
    # Vérifier rôles personnalisés par mots-clés
    role_lower = current_user.role.lower()
    if any(keyword in role_lower for keyword in allowed_custom_keywords):
        return current_user
    
    raise HTTPException(
        status_code=status.HTTP_403_FORBIDDEN,
        detail="Accès refusé. Permissions d'inspection requises (chefs de section et supérieurs)."
    )

# Email service (simplified - in production use proper email service)
async def send_invitation_email(email: str, nom: str, prenom: str, token: str):
    # In production, use proper email service like SendGrid, AWS SES, etc.
    # For now, we'll just log the invitation details
    invitation_link = f"http://localhost:3000/invitation?token={token}"
    logging.info(f"INVITATION EMAIL - Destinataire: {email}")
    logging.info(f"Nom: {prenom} {nom}")
    logging.info(f"Lien d'invitation: {invitation_link}")
    logging.info("Note: Intégrer un vrai service d'email pour la production")

# Routes d'authentification
@api_router.post("/auth/login", response_model=Token)
async def login(request: LoginRequest):
    # Trouver l'utilisateur par username
    user_data = await db.users.find_one({"username": request.username})
    if not user_data or not user_data.get("hashed_password"):
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Nom d'utilisateur ou mot de passe incorrect"
        )
    
    # Vérifier le mot de passe
    if not verify_password(request.password, user_data["hashed_password"]):
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Nom d'utilisateur ou mot de passe incorrect"
        )
    
    # Créer le token d'accès
    access_token_expires = timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES)
    access_token = create_access_token(
        data={"sub": user_data["id"]}, expires_delta=access_token_expires
    )
    
    user = User(**user_data)
    return {
        "access_token": access_token,
        "token_type": "bearer",
        "user": user
    }

@api_router.get("/version-info")
async def get_version_info():
    """
    Endpoint public pour vérifier les informations de version
    Accessible sans authentification
    """
    settings_doc = await db.settings.find_one({"type": "app_settings"})
    
    if not settings_doc:
        # Retourner les valeurs par défaut si aucun paramètre n'existe
        return {
            "currentApkVersion": "1.0.0",
            "minimumSupportedVersion": "1.0.0",
            "apkDownloadUrl": "",
            "forceUpdate": False,
            "releaseNotes": []
        }
    
    return {
        "currentApkVersion": settings_doc.get("currentApkVersion", "1.0.0"),
        "minimumSupportedVersion": settings_doc.get("minimumSupportedVersion", "1.0.0"),
        "apkDownloadUrl": settings_doc.get("apkDownloadUrl", ""),
        "forceUpdate": settings_doc.get("forceUpdate", False),
        "releaseNotes": settings_doc.get("releaseNotes", [])
    }

@api_router.post("/auth/invite")
async def invite_user(
    invitation: UserInvitation,
    current_user: User = Depends(require_admin_or_encadrement)
):
    # Si email fourni, vérifier qu'il n'existe pas déjà
    if invitation.email:
        existing_user = await db.users.find_one({"email": invitation.email})
        if existing_user:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Un utilisateur avec cet email existe déjà"
            )
    
    # Créer le token d'invitation seulement si email fourni
    invitation_token = None
    invitation_expires = None
    if invitation.email:
        invitation_token = create_invitation_token(invitation.email)
        invitation_expires = datetime.utcnow() + timedelta(days=7)
    
    # Créer l'utilisateur
    user_data = UserInDB(
        nom=invitation.nom,
        prenom=invitation.prenom,
        email=invitation.email,
        grade=invitation.grade,
        role=invitation.role,
        section_id=invitation.section_id,
        has_admin_privileges=invitation.has_admin_privileges,
        invitation_token=invitation_token,
        invitation_expires=invitation_expires,
        created_by=current_user.id,
        actif=invitation.email is None  # Actif immédiatement si pas d'email
    )
    
    # Convertir datetime en string pour MongoDB
    user_dict = user_data.dict()
    user_dict['created_at'] = user_data.created_at.isoformat()
    if user_data.invitation_expires:
        user_dict['invitation_expires'] = user_data.invitation_expires.isoformat()
    await db.users.insert_one(user_dict)
    
    # Envoyer l'email d'invitation seulement si email fourni
    if invitation.email and invitation_token:
        await send_invitation_email(
            invitation.email, 
            invitation.nom, 
            invitation.prenom, 
            invitation_token
        )
        return {"message": "Invitation envoyée avec succès", "token": invitation_token}
    else:
        return {"message": "Utilisateur créé avec succès (sans email)", "token": None}

@api_router.post("/auth/set-password")
async def set_password(request: SetPasswordRequest):
    try:
        # Décoder le token d'invitation
        payload = jwt.decode(request.token, SECRET_KEY, algorithms=[ALGORITHM])
        email = payload.get("email")
        if not email or payload.get("type") != "invitation":
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Token d'invitation invalide"
            )
    except jwt.ExpiredSignatureError:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Token d'invitation expiré"
        )
    except jwt.PyJWTError:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Token d'invitation invalide"
        )
    
    # Trouver l'utilisateur avec ce token
    user_data = await db.users.find_one({
        "email": email,
        "invitation_token": request.token
    })
    
    if not user_data:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Token d'invitation invalide ou utilisateur non trouvé"
        )
    
    # Vérifier que le token n'a pas expiré
    if user_data["invitation_expires"] < datetime.utcnow():
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Token d'invitation expiré"
        )
    
    # Hasher le mot de passe et activer l'utilisateur
    hashed_password = get_password_hash(request.password)
    
    await db.users.update_one(
        {"email": email},
        {
            "$set": {
                "hashed_password": hashed_password,
                "actif": True
            },
            "$unset": {
                "invitation_token": "",
                "invitation_expires": ""
            }
        }
    )
    
    return {"message": "Mot de passe défini avec succès"}

@api_router.get("/auth/me", response_model=User)
async def get_current_user_info(current_user: User = Depends(get_current_user)):
    return current_user

# Routes pour la gestion des utilisateurs
@api_router.get("/users")
async def get_users(
    grade: Optional[str] = None,
    role: Optional[str] = None,
    section_id: Optional[str] = None,
    current_user: User = Depends(require_inspection_permissions)
):
    """
    Récupérer la liste des utilisateurs - accessible aux inspecteurs
    Filtres optionnels : grade, role, section_id
    """
    try:
        # Construire le filtre de base
        filter_dict = {}
        
        # Ajouter les filtres optionnels
        if grade:
            filter_dict["grade"] = grade
        if role:
            filter_dict["role"] = role
        if section_id:
            filter_dict["section_id"] = section_id
        
        # Récupérer tous les utilisateurs (actifs et en attente) avec les filtres appliqués
        users = await db.users.find(filter_dict).to_list(1000)
        
        # Convert to User models and then to dict for JSON serialization
        user_models = []
        for user in users:
            try:
                user_model = User(**user)
                user_models.append(user_model.dict())
            except Exception as e:
                # Log the error but continue
                print(f"Error converting user {user.get('prenom', 'N/A')} {user.get('nom', 'N/A')}: {e}")
                continue
        
        return user_models
    except Exception as e:
        print(f"Error in get_users endpoint: {e}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Internal server error: {str(e)}")

@api_router.post("/users", response_model=dict)
async def create_user(
    user: UserCreate, 
    current_user: User = Depends(require_admin_or_encadrement)
):
    # Vérifier si l'utilisateur existe déjà par email
    if user.email:
        existing_user = await db.users.find_one({"email": user.email})
        if existing_user:
            raise HTTPException(
                status_code=status.HTTP_409_CONFLICT,
                detail="Un utilisateur avec cet email existe déjà"
            )
    
    # Générer un username unique automatiquement
    username = await generate_unique_username(user.prenom, user.nom)
    
    # Créer l'utilisateur - TOUJOURS ACTIF lors de la création par admin
    new_user = {
        "id": str(uuid.uuid4()),
        "prenom": user.prenom,
        "nom": user.nom,
        "username": username,  # Username généré automatiquement
        "email": user.email,
        "password_hash": None,  # Pas de mot de passe initial
        "role": user.role,
        "grade": user.grade,
        "section_id": user.section_id,
        "subgroup_id": user.subgroup_id,  # Sous-groupe optionnel
        "actif": True,  # 🔥 TOUJOURS ACTIF lors de la création par admin
        "has_admin_privileges": user.has_admin_privileges,
        "created_at": datetime.utcnow()
    }
    
    await db.users.insert_one(new_user)
    
    # Envoyer l'invitation par email si un email est fourni
    if user.email:
        try:
            # Ici on pourrait envoyer un vrai email d'invitation
            # Pour l'instant, on simule l'envoi
            print(f"📧 Email d'invitation envoyé à {user.email}")
        except Exception as e:
            print(f"⚠️ Erreur lors de l'envoi de l'email: {e}")
    
    return {
        "message": "Utilisateur créé avec succès", 
        "user_id": new_user["id"],
        "username": username  # Retourner le username généré
    }

@api_router.get("/users/filters")
async def get_user_filters(current_user: User = Depends(require_admin_or_encadrement)):
    """Récupérer les options de filtres pour les utilisateurs"""
    
    # Récupérer tous les utilisateurs pour extraire les filtres
    users = await db.users.find({"actif": True}).to_list(1000)
    sections = await db.sections.find().to_list(1000)
    
    # Extraire les grades uniques
    grades = list(set([user.get("grade") for user in users if user.get("grade")]))
    grades.sort()
    
    # Extraire les rôles uniques
    roles = list(set([user.get("role") for user in users if user.get("role")]))
    roles.sort()
    
    # Formatter les sections
    section_options = [{"id": section["id"], "name": section["nom"]} for section in sections]
    section_options.sort(key=lambda x: x["name"])
    
    return {
        "grades": grades,
        "roles": roles,
        "sections": section_options
    }

@api_router.get("/users/{user_id}", response_model=User)
async def get_user(user_id: str, current_user: User = Depends(get_current_user)):
    # Les utilisateurs peuvent voir leur propre profil
    # Les admins/encadrement peuvent voir tous les profils
    if (current_user.id != user_id and 
        current_user.role not in [UserRole.CADET_ADMIN, UserRole.ENCADREMENT]):
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Accès refusé"
        )
    
    user = await db.users.find_one({"id": user_id, "actif": True})
    if not user:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Utilisateur non trouvé"
        )
    return User(**user)

@api_router.put("/users/{user_id}")
async def update_user(
    user_id: str,
    user_update: UserUpdate,
    current_user: User = Depends(require_admin_or_encadrement)
):
    # Vérifier que l'utilisateur existe
    existing_user = await db.users.find_one({"id": user_id, "actif": True})
    if not existing_user:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Utilisateur non trouvé"
        )
    
    # Préparer les mises à jour (seulement les champs fournis)
    update_data = {}
    
    if user_update.nom is not None:
        update_data["nom"] = user_update.nom.strip()
    if user_update.prenom is not None:
        update_data["prenom"] = user_update.prenom.strip()
    if user_update.email is not None:
        # Vérifier que l'email n'est pas déjà utilisé par un autre utilisateur
        if user_update.email:
            existing_email = await db.users.find_one({
                "email": user_update.email,
                "id": {"$ne": user_id}
            })
            if existing_email:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail="Cet email est déjà utilisé par un autre utilisateur"
                )
            
            # Si l'utilisateur n'avait pas d'email avant et qu'on lui en ajoute un
            if not existing_user.get("email") and user_update.email:
                # Créer un token d'invitation et l'envoyer
                invitation_token = create_invitation_token(user_update.email)
                update_data["invitation_token"] = invitation_token
                update_data["invitation_expires"] = (datetime.utcnow() + timedelta(days=7)).isoformat()
                update_data["actif"] = False  # L'utilisateur devra confirmer par email
                
                # Envoyer l'email d'invitation
                await send_invitation_email(
                    user_update.email,
                    existing_user["nom"],
                    existing_user["prenom"],
                    invitation_token
                )
        
        update_data["email"] = user_update.email
    if user_update.grade is not None:
        update_data["grade"] = user_update.grade.value
    if user_update.role is not None:
        update_data["role"] = user_update.role  # Plus besoin de .value car c'est maintenant un str
    if user_update.section_id is not None:
        # Vérifier que la section existe si fournie
        if user_update.section_id:
            section = await db.sections.find_one({"id": user_update.section_id})
            if not section:
                raise HTTPException(
                    status_code=status.HTTP_404_NOT_FOUND,
                    detail="Section non trouvée"
                )
        update_data["section_id"] = user_update.section_id
    if user_update.subgroup_id is not None:
        # Vérifier que le sous-groupe existe si fourni
        if user_update.subgroup_id:
            subgroup = await db.subgroups.find_one({"id": user_update.subgroup_id})
            if not subgroup:
                raise HTTPException(
                    status_code=status.HTTP_404_NOT_FOUND,
                    detail="Sous-groupe non trouvé"
                )
            # Vérifier que le sous-groupe appartient à la section de l'utilisateur
            if update_data.get("section_id") or existing_user.get("section_id"):
                expected_section_id = update_data.get("section_id") or existing_user.get("section_id")
                if subgroup["section_id"] != expected_section_id:
                    raise HTTPException(
                        status_code=status.HTTP_400_BAD_REQUEST,
                        detail="Le sous-groupe doit appartenir à la section de l'utilisateur"
                    )
        update_data["subgroup_id"] = user_update.subgroup_id
    if user_update.actif is not None:
        update_data["actif"] = user_update.actif
    if user_update.has_admin_privileges is not None:
        update_data["has_admin_privileges"] = user_update.has_admin_privileges
    
    # Effectuer la mise à jour
    if update_data:
        await db.users.update_one(
            {"id": user_id},
            {"$set": update_data}
        )
    
    return {"message": "Utilisateur mis à jour avec succès"}

@api_router.delete("/users/{user_id}")
async def delete_user(
    user_id: str,
    current_user: User = Depends(require_admin_or_encadrement)
):
    # Vérifier que l'utilisateur existe
    existing_user = await db.users.find_one({"id": user_id, "actif": True})
    if not existing_user:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Utilisateur non trouvé"
        )
    
    # Empêcher la suppression de son propre compte
    if user_id == current_user.id:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Vous ne pouvez pas supprimer votre propre compte"
        )
    
    # Supprimer définitivement l'utilisateur et toutes ses données associées
    try:
        # Supprimer toutes les présences de cet utilisateur
        await db.presences.delete_many({"cadet_id": user_id})
        
        # Supprimer l'utilisateur des activités
        await db.activities.update_many(
            {"cadet_ids": user_id},
            {"$pull": {"cadet_ids": user_id}}
        )
        
        # Supprimer l'utilisateur
        result = await db.users.delete_one({"id": user_id})
        
        if result.deleted_count == 0:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Utilisateur non trouvé"
            )
        
        return {"message": f"Utilisateur {existing_user['prenom']} {existing_user['nom']} supprimé définitivement avec toutes ses données"}
    
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Erreur lors de la suppression: {str(e)}"
        )

@api_router.post("/users/{user_id}/generate-password", response_model=GeneratePasswordResponse)
async def generate_initial_password(
    user_id: str,
    current_user: User = Depends(require_admin_or_encadrement)
):
    """
    Génère un mot de passe initial aléatoire pour un utilisateur
    L'utilisateur devra changer ce mot de passe à sa première connexion
    """
    # Vérifier que l'utilisateur existe
    existing_user = await db.users.find_one({"id": user_id})
    if not existing_user:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Utilisateur non trouvé"
        )
    
    # Générer un username si l'utilisateur n'en a pas
    username = existing_user.get("username")
    if not username:
        import re
        prenom = existing_user.get("prenom", "")
        nom = existing_user.get("nom", "")
        
        # Générer username : prenom.nom en minuscules sans accents
        def remove_accents(text):
            import unicodedata
            return ''.join(c for c in unicodedata.normalize('NFD', text) if unicodedata.category(c) != 'Mn')
        
        username_base = f"{remove_accents(prenom).lower()}.{remove_accents(nom).lower()}"
        username_base = re.sub(r'[^a-z.]', '', username_base)
        
        # Vérifier si le username existe déjà
        existing_username = await db.users.find_one({"username": username_base})
        if existing_username:
            # Ajouter un numéro
            counter = 2
            while await db.users.find_one({"username": f"{username_base}{counter}"}):
                counter += 1
            username = f"{username_base}{counter}"
        else:
            username = username_base
        
        # Mettre à jour l'utilisateur avec le nouveau username
        await db.users.update_one(
            {"id": user_id},
            {"$set": {"username": username}}
        )
    
    # Générer un mot de passe aléatoire de 8 caractères
    import random
    import string
    characters = string.ascii_letters + string.digits
    temporary_password = ''.join(random.choice(characters) for _ in range(8))
    
    # Hasher le mot de passe
    hashed_password = get_password_hash(temporary_password)
    
    # Mettre à jour l'utilisateur avec le nouveau mot de passe et le flag must_change_password
    await db.users.update_one(
        {"id": user_id},
        {
            "$set": {
                "hashed_password": hashed_password,
                "must_change_password": True,
                "actif": True  # S'assurer que l'utilisateur est actif
            }
        }
    )
    
    return GeneratePasswordResponse(
        user_id=user_id,
        username=username,
        temporary_password=temporary_password,
        message="Mot de passe temporaire généré avec succès. L'utilisateur devra le changer à sa première connexion."
    )

@api_router.post("/auth/change-password")
async def change_password(
    request: ChangePasswordRequest,
    current_user: User = Depends(get_current_user)
):
    """
    Permet à un utilisateur de changer son mot de passe
    Nécessite l'ancien mot de passe pour validation
    """
    # Récupérer l'utilisateur avec son mot de passe
    user_data = await db.users.find_one({"id": current_user.id})
    if not user_data or not user_data.get("hashed_password"):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Aucun mot de passe défini pour cet utilisateur"
        )
    
    # Vérifier l'ancien mot de passe
    if not verify_password(request.old_password, user_data["hashed_password"]):
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Mot de passe actuel incorrect"
        )
    
    # Valider le nouveau mot de passe
    if len(request.new_password) < 6:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Le nouveau mot de passe doit contenir au moins 6 caractères"
        )
    
    # Hasher le nouveau mot de passe
    new_hashed_password = get_password_hash(request.new_password)
    
    # Mettre à jour le mot de passe et retirer le flag must_change_password
    await db.users.update_one(
        {"id": current_user.id},
        {
            "$set": {
                "hashed_password": new_hashed_password,
                "must_change_password": False
            }
        }
    )
    
    return {"message": "Mot de passe changé avec succès"}

@api_router.get("/auth/profile", response_model=User)
async def get_profile(current_user: User = Depends(get_current_user)):
    """
    Retourne les informations du profil de l'utilisateur connecté
    """
    # Récupérer les informations complètes depuis la base de données
    user_data = await db.users.find_one({"id": current_user.id})
    if not user_data:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Utilisateur non trouvé"
        )
    
    return User(**user_data)

# Routes pour les sections
@api_router.post("/sections", response_model=Section)
async def create_section(
    section: SectionCreate,
    current_user: User = Depends(require_admin_or_encadrement)
):
    section_data = Section(**section.dict())
    # Convertir datetime en string pour MongoDB
    section_dict = section_data.dict()
    section_dict['created_at'] = section_data.created_at.isoformat()
    await db.sections.insert_one(section_dict)
    return section_data

@api_router.get("/sections", response_model=List[Section])
async def get_sections(current_user: User = Depends(get_current_user)):
    sections = await db.sections.find().to_list(1000)
    return [Section(**section) for section in sections]

@api_router.put("/sections/{section_id}")
async def update_section(
    section_id: str,
    section_update: SectionCreate,
    current_user: User = Depends(require_admin_or_encadrement)
):
    # Vérifier que la section existe
    existing_section = await db.sections.find_one({"id": section_id})
    if not existing_section:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Section non trouvée"
        )
    
    # Préparer les données de mise à jour
    update_data = {
        "nom": section_update.nom,
        "description": section_update.description,
        "responsable_id": section_update.responsable_id if section_update.responsable_id else None
    }
    
    # Mettre à jour la section
    await db.sections.update_one(
        {"id": section_id},
        {"$set": update_data}
    )
    
    return {"message": "Section mise à jour avec succès"}

@api_router.delete("/sections/{section_id}")
async def delete_section(
    section_id: str,
    current_user: User = Depends(require_admin_or_encadrement)
):
    # Vérifier que la section existe
    existing_section = await db.sections.find_one({"id": section_id})
    if not existing_section:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Section non trouvée"
        )
    
    # Supprimer définitivement la section et mettre à jour les utilisateurs
    try:
        # Retirer l'affectation de section de tous les utilisateurs
        await db.users.update_many(
            {"section_id": section_id},
            {"$unset": {"section_id": ""}}
        )
        
        # Supprimer la section
        result = await db.sections.delete_one({"id": section_id})
        
        if result.deleted_count == 0:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Section non trouvée"
            )
        
        return {"message": f"Section {existing_section['nom']} supprimée définitivement"}
    
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Erreur lors de la suppression: {str(e)}"
        )

# Routes pour les sous-groupes
@api_router.get("/sections/{section_id}/subgroups", response_model=List[SubGroup])
async def get_subgroups(
    section_id: str,
    current_user: User = Depends(require_admin_or_encadrement)
):
    """Récupérer tous les sous-groupes d'une section"""
    # Vérifier que la section existe
    section = await db.sections.find_one({"id": section_id})
    if not section:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Section non trouvée"
        )
    
    subgroups = await db.subgroups.find({"section_id": section_id}).to_list(100)
    return [SubGroup(**subgroup) for subgroup in subgroups]

@api_router.post("/subgroups", response_model=SubGroup)
async def create_subgroup(
    subgroup: SubGroupCreate,
    current_user: User = Depends(require_admin_or_encadrement)
):
    """Créer un nouveau sous-groupe"""
    # Vérifier que la section existe
    section = await db.sections.find_one({"id": subgroup.section_id})
    if not section:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Section non trouvée"
        )
    
    # Vérifier que le nom du sous-groupe n'existe pas déjà dans cette section
    existing_subgroup = await db.subgroups.find_one({
        "nom": subgroup.nom,
        "section_id": subgroup.section_id
    })
    if existing_subgroup:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Un sous-groupe avec ce nom existe déjà dans cette section"
        )
    
    # Vérifier que le responsable existe si fourni
    if subgroup.responsable_id:
        responsable = await db.users.find_one({"id": subgroup.responsable_id, "actif": True})
        if not responsable:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Responsable non trouvé ou inactif"
            )
    
    new_subgroup = SubGroup(**subgroup.dict())
    # Convertir datetime en string pour MongoDB
    subgroup_dict = new_subgroup.dict()
    subgroup_dict['created_at'] = new_subgroup.created_at.isoformat()
    await db.subgroups.insert_one(subgroup_dict)
    
    return new_subgroup

@api_router.put("/subgroups/{subgroup_id}")
async def update_subgroup(
    subgroup_id: str,
    subgroup_update: SubGroupUpdate,
    current_user: User = Depends(require_admin_or_encadrement)
):
    """Mettre à jour un sous-groupe"""
    # Vérifier que le sous-groupe existe
    existing_subgroup = await db.subgroups.find_one({"id": subgroup_id})
    if not existing_subgroup:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Sous-groupe non trouvé"
        )
    
    # Préparer les mises à jour
    update_data = {}
    
    if subgroup_update.nom is not None:
        # Vérifier que le nouveau nom n'existe pas déjà dans la section
        name_exists = await db.subgroups.find_one({
            "nom": subgroup_update.nom,
            "section_id": existing_subgroup["section_id"],
            "id": {"$ne": subgroup_id}
        })
        if name_exists:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Un sous-groupe avec ce nom existe déjà dans cette section"
            )
        update_data["nom"] = subgroup_update.nom
    
    if subgroup_update.description is not None:
        update_data["description"] = subgroup_update.description
    
    if subgroup_update.responsable_id is not None:
        # Vérifier que le responsable existe
        if subgroup_update.responsable_id:  # Si ce n'est pas une chaîne vide
            responsable = await db.users.find_one({"id": subgroup_update.responsable_id, "actif": True})
            if not responsable:
                raise HTTPException(
                    status_code=status.HTTP_404_NOT_FOUND,
                    detail="Responsable non trouvé ou inactif"
                )
        update_data["responsable_id"] = subgroup_update.responsable_id
    
    # Effectuer la mise à jour
    if update_data:
        await db.subgroups.update_one(
            {"id": subgroup_id},
            {"$set": update_data}
        )
    
    return {"message": "Sous-groupe mis à jour avec succès"}

@api_router.delete("/subgroups/{subgroup_id}")
async def delete_subgroup(
    subgroup_id: str,
    current_user: User = Depends(require_admin_or_encadrement)
):
    """Supprimer un sous-groupe"""
    # Vérifier que le sous-groupe existe
    existing_subgroup = await db.subgroups.find_one({"id": subgroup_id})
    if not existing_subgroup:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Sous-groupe non trouvé"
        )
    
    # Supprimer le sous-groupe et mettre à jour les utilisateurs
    try:
        # Retirer l'affectation de sous-groupe de tous les utilisateurs
        await db.users.update_many(
            {"subgroup_id": subgroup_id},
            {"$unset": {"subgroup_id": ""}}
        )
        
        # Supprimer le sous-groupe
        result = await db.subgroups.delete_one({"id": subgroup_id})
        
        if result.deleted_count == 0:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Sous-groupe non trouvé"
            )
        
        return {"message": f"Sous-groupe {existing_subgroup['nom']} supprimé définitivement"}
    
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Erreur lors de la suppression: {str(e)}"
        )

# Routes pour les présences
@api_router.post("/presences", response_model=Presence)
async def create_presence(
    presence: PresenceCreate,
    presence_date: date = None,
    activite: Optional[str] = None,
    current_user: User = Depends(require_presence_permissions)
):
    # Utiliser la date d'aujourd'hui si non fournie
    if presence_date is None:
        presence_date = date.today()
    
    # Gérer les invités différemment
    if presence.is_guest:
        # Valider que les informations de l'invité sont présentes
        if not presence.guest_nom or not presence.guest_prenom:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Le nom et le prénom de l'invité sont requis"
            )
        
        # Générer un ID unique pour l'invité
        guest_id = f"guest_{str(uuid.uuid4())}"
        
        # Créer la présence pour l'invité
        presence_data = Presence(
            cadet_id=guest_id,
            date=presence_date,
            status=presence.status,
            commentaire=presence.commentaire,
            enregistre_par=current_user.id,
            section_id=current_user.section_id,  # Assigner à la section de celui qui enregistre
            activite=activite,
            is_guest=True,
            guest_nom=presence.guest_nom,
            guest_prenom=presence.guest_prenom
        )
    else:
        # Logique existante pour les cadets réguliers
        # Vérifier que le cadet existe
        cadet = await db.users.find_one({"id": presence.cadet_id, "actif": True})
        if not cadet:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Cadet non trouvé"
            )
        
        # Vérifier les permissions selon le rôle
        if current_user.role == UserRole.CADET_RESPONSIBLE:
            # Un cadet responsable ne peut enregistrer que pour sa section
            if cadet.get("section_id") != current_user.section_id:
                raise HTTPException(
                    status_code=status.HTTP_403_FORBIDDEN,
                    detail="Vous ne pouvez enregistrer les présences que pour votre section"
                )
        
        # Vérifier si une présence existe déjà pour ce cadet à cette date
        existing_presence = await db.presences.find_one({
            "cadet_id": presence.cadet_id,
            "date": presence_date.isoformat()
        })
        
        if existing_presence:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Une présence existe déjà pour ce cadet à cette date"
            )
        
        # Créer la présence
        presence_data = Presence(
            cadet_id=presence.cadet_id,
            date=presence_date,
            status=presence.status,
            commentaire=presence.commentaire,
            enregistre_par=current_user.id,
            section_id=cadet.get("section_id"),
            activite=activite,
            is_guest=False
        )
    
    # Enregistrer dans MongoDB (convertir date et datetime en string)
    presence_dict = presence_data.dict()
    presence_dict['date'] = presence_data.date.isoformat()
    presence_dict['heure_enregistrement'] = presence_data.heure_enregistrement.isoformat()
    await db.presences.insert_one(presence_dict)
    
    return presence_data

@api_router.post("/presences/bulk")
async def create_bulk_presences(
    bulk_data: PresenceBulkCreate,
    current_user: User = Depends(require_presence_permissions)
):
    created_presences = []
    errors = []
    
    for presence_create in bulk_data.presences:
        try:
            # Vérifier que le cadet existe
            cadet = await db.users.find_one({"id": presence_create.cadet_id, "actif": True})
            if not cadet:
                errors.append(f"Cadet {presence_create.cadet_id} non trouvé")
                continue
            
            # Vérifier les permissions selon le rôle
            if current_user.role == UserRole.CADET_RESPONSIBLE:
                if cadet.get("section_id") != current_user.section_id:
                    errors.append(f"Permission refusée pour le cadet {cadet['prenom']} {cadet['nom']}")
                    continue
            
            # Vérifier si une présence existe déjà
            existing_presence = await db.presences.find_one({
                "cadet_id": presence_create.cadet_id,
                "date": bulk_data.date.isoformat()
            })
            
            if existing_presence:
                # Mettre à jour la présence existante
                await db.presences.update_one(
                    {"id": existing_presence["id"]},
                    {"$set": {
                        "status": presence_create.status.value,
                        "commentaire": presence_create.commentaire,
                        "enregistre_par": current_user.id,
                        "heure_enregistrement": datetime.utcnow().isoformat(),
                        "activite": bulk_data.activite
                    }}
                )
                created_presences.append(existing_presence["id"])
            else:
                # Créer nouvelle présence
                presence_data = {
                    "id": str(uuid.uuid4()),
                    "cadet_id": presence_create.cadet_id,
                    "date": bulk_data.date.isoformat(),
                    "status": presence_create.status.value,
                    "commentaire": presence_create.commentaire,
                    "enregistre_par": current_user.id,
                    "heure_enregistrement": datetime.utcnow().isoformat(),
                    "section_id": cadet.get("section_id"),
                    "activite": bulk_data.activite
                }
                
                await db.presences.insert_one(presence_data)
                created_presences.append(presence_data["id"])
                
        except Exception as e:
            errors.append(f"Erreur pour cadet {presence_create.cadet_id}: {str(e)}")
    
    return {
        "created_count": len(created_presences),
        "created_ids": created_presences,
        "errors": errors
    }

@api_router.get("/presences", response_model=List[PresenceResponse])
async def get_presences(
    date: Optional[date] = None,
    cadet_id: Optional[str] = None,
    section_id: Optional[str] = None,
    limit: int = 100,
    current_user: User = Depends(get_current_user)
):
    # Construire le filtre selon les permissions
    filter_dict = {}
    
    if current_user.role == UserRole.CADET:
        # Un cadet ne peut voir que ses propres présences
        filter_dict["cadet_id"] = current_user.id
    elif current_user.role == UserRole.CADET_RESPONSIBLE:
        # Un cadet responsable ne peut voir que sa section
        if current_user.section_id:
            filter_dict["section_id"] = current_user.section_id
        else:
            # Si pas de section assignée, ne peut rien voir
            return []
    # CADET_ADMIN et ENCADREMENT peuvent tout voir
    
    # Appliquer les filtres additionnels
    if date:
        filter_dict["date"] = date.isoformat()
    if cadet_id and current_user.role in [UserRole.CADET_ADMIN, UserRole.ENCADREMENT]:
        filter_dict["cadet_id"] = cadet_id
    if section_id and current_user.role in [UserRole.CADET_ADMIN, UserRole.ENCADREMENT]:
        filter_dict["section_id"] = section_id
    
    # Récupérer les présences
    presences_cursor = db.presences.find(filter_dict).limit(limit).sort("date", -1)
    presences = await presences_cursor.to_list(limit)
    
    # Enrichir avec les informations des cadets et sections
    enriched_presences = []
    for presence in presences:
        # Gérer les invités différemment
        if presence.get("is_guest", False):
            # Pour les invités, utiliser les infos stockées
            cadet_nom = presence.get("guest_nom", "Inconnu")
            cadet_prenom = presence.get("guest_prenom", "Inconnu")
        else:
            # Récupérer les infos du cadet
            cadet = await db.users.find_one({"id": presence["cadet_id"]})
            if not cadet:
                continue
            cadet_nom = cadet["nom"]
            cadet_prenom = cadet["prenom"]
            
        # Récupérer les infos de la section si applicable
        section_nom = None
        if presence.get("section_id"):
            section = await db.sections.find_one({"id": presence["section_id"]})
            if section:
                section_nom = section["nom"]
        
        enriched_presence = PresenceResponse(
            id=presence["id"],
            cadet_id=presence["cadet_id"],
            cadet_nom=cadet_nom,
            cadet_prenom=cadet_prenom,
            date=datetime.fromisoformat(presence["date"]).date() if isinstance(presence["date"], str) else presence["date"],
            status=PresenceStatus(presence["status"]),
            commentaire=presence.get("commentaire"),
            enregistre_par=presence["enregistre_par"],
            heure_enregistrement=datetime.fromisoformat(presence["heure_enregistrement"]) if isinstance(presence["heure_enregistrement"], str) else presence["heure_enregistrement"],
            section_id=presence.get("section_id"),
            section_nom=section_nom,
            activite=presence.get("activite"),
            is_guest=presence.get("is_guest", False),
            guest_nom=presence.get("guest_nom"),
            guest_prenom=presence.get("guest_prenom")
        )
        enriched_presences.append(enriched_presence)
    
    return enriched_presences

@api_router.put("/presences/{presence_id}")
async def update_presence(
    presence_id: str,
    updates: PresenceUpdate,
    current_user: User = Depends(require_presence_permissions)
):
    # Vérifier que la présence existe
    presence = await db.presences.find_one({"id": presence_id})
    if not presence:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Présence non trouvée"
        )
    
    # Vérifier les permissions
    if current_user.role == UserRole.CADET_RESPONSIBLE:
        if presence.get("section_id") != current_user.section_id:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="Vous ne pouvez modifier que les présences de votre section"
            )
    
    # Vérifier le délai de 24h (seulement pour les non-admins)
    if current_user.role not in [UserRole.CADET_ADMIN, UserRole.ENCADREMENT]:
        # Récupérer la date de la présence
        presence_datetime = presence.get("heure_enregistrement")
        if isinstance(presence_datetime, str):
            presence_datetime = datetime.fromisoformat(presence_datetime.replace('Z', '+00:00'))
        
        # Calculer le délai
        now = datetime.now(timezone.utc)
        if presence_datetime.tzinfo is None:
            presence_datetime = presence_datetime.replace(tzinfo=timezone.utc)
        
        time_diff = now - presence_datetime
        
        # Si plus de 24h, interdire la modification
        if time_diff.total_seconds() > 86400:  # 24h = 86400 secondes
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="Vous ne pouvez modifier une présence que dans les 24h suivant son enregistrement. Contactez un administrateur."
            )
    
    # Préparer les mises à jour
    update_data = {}
    if updates.status is not None:
        update_data["status"] = updates.status.value
    if updates.commentaire is not None:
        update_data["commentaire"] = updates.commentaire
    
    update_data["enregistre_par"] = current_user.id
    update_data["heure_enregistrement"] = datetime.utcnow().isoformat()
    
    # Mettre à jour
    await db.presences.update_one(
        {"id": presence_id},
        {"$set": update_data}
    )
    
    return {"message": "Présence mise à jour avec succès"}

@api_router.get("/presences/stats/{cadet_id}", response_model=PresenceStats)
async def get_presence_stats(
    cadet_id: str,
    date_debut: Optional[date] = None,
    date_fin: Optional[date] = None,
    current_user: User = Depends(get_current_user)
):
    # Vérifier les permissions
    if current_user.role == UserRole.CADET and current_user.id != cadet_id:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Vous ne pouvez consulter que vos propres statistiques"
        )
    
    # Construire le filtre
    filter_dict = {"cadet_id": cadet_id}
    if date_debut:
        filter_dict["date"] = {"$gte": date_debut.isoformat()}
    if date_fin:
        if "date" in filter_dict:
            filter_dict["date"]["$lte"] = date_fin.isoformat()
        else:
            filter_dict["date"] = {"$lte": date_fin.isoformat()}
    
    # Récupérer toutes les présences
    presences = await db.presences.find(filter_dict).to_list(1000)
    
    # Calculer les statistiques
    total_seances = len(presences)
    presences_count = len([p for p in presences if p["status"] == "present"])
    absences = len([p for p in presences if p["status"] == "absent"])
    retards = len([p for p in presences if p["status"] == "retard"])
    
    taux_presence = (presences_count / total_seances * 100) if total_seances > 0 else 0
    
    return PresenceStats(
        total_seances=total_seances,
        presences=presences_count,
        absences=absences,
        absences_excusees=0,  # Maintenu à 0 pour compatibilité
        retards=retards,
        taux_presence=round(taux_presence, 2)
    )

# Routes pour les activités pré-définies
@api_router.post("/activities", response_model=Activity)
async def create_activity(
    activity: ActivityCreate,
    current_user: User = Depends(require_admin_or_encadrement)
):
    # Vérifier que tous les cadets existent (peu importe leur statut d'activation)
    for cadet_id in activity.cadet_ids:
        cadet = await db.users.find_one({"id": cadet_id})
        if not cadet:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail=f"Cadet {cadet_id} non trouvé"
            )
    
    # Créer l'activité
    activity_data = Activity(
        nom=activity.nom,
        description=activity.description,
        type=activity.type,
        cadet_ids=activity.cadet_ids,
        recurrence_interval=activity.recurrence_interval,
        recurrence_unit=activity.recurrence_unit,
        next_date=activity.next_date,
        planned_date=activity.planned_date,
        created_by=current_user.id
    )
    
    # Convertir datetime en string pour MongoDB
    activity_dict = activity_data.dict()
    activity_dict['created_at'] = activity_data.created_at.isoformat()
    await db.activities.insert_one(activity_dict)
    return activity_data

@api_router.get("/activities", response_model=List[ActivityResponse])
async def get_activities(
    active_only: bool = True,
    current_user: User = Depends(require_admin_or_encadrement)
):
    filter_dict = {}
    if active_only:
        filter_dict["active"] = True
    
    activities = await db.activities.find(filter_dict).to_list(1000)
    
    # Enrichir avec les noms des cadets
    enriched_activities = []
    for activity in activities:
        # Récupérer les noms des cadets (actifs et non actifs)
        cadet_names = []
        for cadet_id in activity["cadet_ids"]:
            cadet = await db.users.find_one({"id": cadet_id})
            if cadet:
                status_indicator = "" if cadet.get("actif", False) else " (non confirmé)"
                cadet_names.append(f"{cadet['prenom']} {cadet['nom']}{status_indicator}")
        
        # Conversion des dates pour la réponse
        next_date = None
        if activity.get("next_date"):
            if isinstance(activity["next_date"], str):
                next_date = activity["next_date"]
            else:
                # Conversion date object vers string
                next_date = activity["next_date"].strftime("%Y-%m-%d")
        
        planned_date = None
        if activity.get("planned_date"):
            if isinstance(activity["planned_date"], str):
                planned_date = activity["planned_date"]
            else:
                # Conversion date object vers string
                planned_date = activity["planned_date"].strftime("%Y-%m-%d")

        # Conversion de created_at
        created_at = activity["created_at"]
        if isinstance(created_at, str):
            created_at = datetime.fromisoformat(created_at)

        enriched_activity = ActivityResponse(
            id=activity["id"],
            nom=activity["nom"],
            description=activity.get("description"),
            type=ActivityType(activity["type"]),
            cadet_ids=activity["cadet_ids"],
            cadet_names=cadet_names,
            recurrence_interval=activity.get("recurrence_interval"),
            recurrence_unit=activity.get("recurrence_unit"),
            next_date=next_date,
            planned_date=planned_date,
            created_by=activity["created_by"],
            created_at=created_at,
            active=activity["active"]
        )
        enriched_activities.append(enriched_activity)
    
    return enriched_activities

@api_router.get("/activities/{activity_id}", response_model=ActivityResponse)
async def get_activity(
    activity_id: str,
    current_user: User = Depends(require_admin_or_encadrement)
):
    activity = await db.activities.find_one({"id": activity_id})
    if not activity:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Activité non trouvée"
        )
    
    # Récupérer les noms des cadets (actifs et non actifs)
    cadet_names = []
    for cadet_id in activity["cadet_ids"]:
        cadet = await db.users.find_one({"id": cadet_id})
        if cadet:
            status_indicator = "" if cadet.get("actif", False) else " (non confirmé)"
            cadet_names.append(f"{cadet['prenom']} {cadet['nom']}{status_indicator}")
    
    # Conversion des dates pour la réponse
    next_date = None
    if activity.get("next_date"):
        if isinstance(activity["next_date"], str):
            next_date = activity["next_date"]
        else:
            # Conversion date object vers string
            next_date = activity["next_date"].strftime("%Y-%m-%d")
    
    planned_date = None
    if activity.get("planned_date"):
        if isinstance(activity["planned_date"], str):
            planned_date = activity["planned_date"]
        else:
            # Conversion date object vers string
            planned_date = activity["planned_date"].strftime("%Y-%m-%d")

    # Conversion de created_at
    created_at = activity["created_at"]
    if isinstance(created_at, str):
        created_at = datetime.fromisoformat(created_at)

    return ActivityResponse(
        id=activity["id"],
        nom=activity["nom"],
        description=activity.get("description"),
        type=ActivityType(activity["type"]),
        cadet_ids=activity["cadet_ids"],
        cadet_names=cadet_names,
        recurrence_interval=activity.get("recurrence_interval"),
        recurrence_unit=activity.get("recurrence_unit"),
        next_date=next_date,
        planned_date=planned_date,
        created_by=activity["created_by"],
        created_at=created_at,
        active=activity["active"]
    )

@api_router.put("/activities/{activity_id}")
async def update_activity(
    activity_id: str,
    activity_update: ActivityCreate,
    current_user: User = Depends(require_admin_or_encadrement)
):
    # Vérifier que l'activité existe
    existing_activity = await db.activities.find_one({"id": activity_id})
    if not existing_activity:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Activité non trouvée"
        )
    
    # Vérifier que tous les cadets existent
    for cadet_id in activity_update.cadet_ids:
        cadet = await db.users.find_one({"id": cadet_id, "actif": True})
        if not cadet:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail=f"Cadet {cadet_id} non trouvé"
            )
    
    # Mettre à jour
    update_data = {
        "nom": activity_update.nom,
        "description": activity_update.description,
        "type": activity_update.type.value,
        "cadet_ids": activity_update.cadet_ids,
        "recurrence_interval": activity_update.recurrence_interval,
        "recurrence_unit": activity_update.recurrence_unit,
        "next_date": activity_update.next_date if activity_update.next_date else None,
        "planned_date": activity_update.planned_date if activity_update.planned_date else None
    }
    
    await db.activities.update_one(
        {"id": activity_id},
        {"$set": update_data}
    )
    
    return {"message": "Activité mise à jour avec succès"}

@api_router.delete("/activities/{activity_id}")
async def delete_activity(
    activity_id: str,
    current_user: User = Depends(require_admin_or_encadrement)
):
    result = await db.activities.update_one(
        {"id": activity_id},
        {"$set": {"active": False}}
    )
    
    if result.matched_count == 0:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Activité non trouvée"
        )
    
    return {"message": "Activité désactivée avec succès"}

# Routes pour la gestion des rôles
@api_router.get("/roles", response_model=List[Role])
async def get_roles(current_user: User = Depends(require_admin_or_encadrement)):
    """Récupérer tous les rôles"""
    roles = await db.roles.find().to_list(1000)
    return [Role(**role) for role in roles]

@api_router.post("/roles", response_model=Role)
async def create_role(
    role: RoleCreate,
    current_user: User = Depends(require_admin_or_encadrement)
):
    """Créer un nouveau rôle"""
    # Vérifier que le nom du rôle n'existe pas déjà
    existing_role = await db.roles.find_one({"name": role.name})
    if existing_role:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Un rôle avec ce nom existe déjà"
        )
    
    role_data = Role(**role.dict())
    # Convertir datetime en string pour MongoDB
    role_dict = role_data.dict()
    role_dict['created_at'] = role_data.created_at.isoformat()
    await db.roles.insert_one(role_dict)
    return role_data

@api_router.put("/roles/{role_id}")
async def update_role(
    role_id: str,
    role_update: RoleUpdate,
    current_user: User = Depends(require_admin_or_encadrement)
):
    """Mettre à jour un rôle"""
    # Vérifier que le rôle existe
    existing_role = await db.roles.find_one({"id": role_id})
    if not existing_role:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Rôle non trouvé"
        )
    
    # Empêcher la modification des rôles système
    if existing_role.get("is_system_role", False):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Les rôles système ne peuvent pas être modifiés"
        )
    
    # Préparer les mises à jour
    update_data = {}
    if role_update.name is not None:
        # Vérifier que le nouveau nom n'existe pas déjà
        name_exists = await db.roles.find_one({
            "name": role_update.name,
            "id": {"$ne": role_id}
        })
        if name_exists:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Un rôle avec ce nom existe déjà"
            )
        update_data["name"] = role_update.name
    
    if role_update.description is not None:
        update_data["description"] = role_update.description
    
    if role_update.permissions is not None:
        update_data["permissions"] = [perm.value for perm in role_update.permissions]
    
    # Effectuer la mise à jour
    if update_data:
        await db.roles.update_one(
            {"id": role_id},
            {"$set": update_data}
        )
    
    return {"message": "Rôle mis à jour avec succès"}

@api_router.delete("/roles/{role_id}")
async def delete_role(
    role_id: str,
    current_user: User = Depends(require_admin_or_encadrement)
):
    """Supprimer un rôle"""
    # Vérifier que le rôle existe
    existing_role = await db.roles.find_one({"id": role_id})
    if not existing_role:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Rôle non trouvé"
        )
    
    # Empêcher la suppression des rôles système
    if existing_role.get("is_system_role", False):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Les rôles système ne peuvent pas être supprimés"
        )
    
    # Vérifier qu'aucun utilisateur n'utilise ce rôle
    users_with_role = await db.users.find({"custom_role_id": role_id}).to_list(1)
    if users_with_role:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Impossible de supprimer un rôle utilisé par des utilisateurs"
        )
    
    # Supprimer le rôle
    result = await db.roles.delete_one({"id": role_id})
    
    if result.deleted_count == 0:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Rôle non trouvé"
        )
    
    return {"message": "Rôle supprimé avec succès"}

# Routes pour les alertes d'absences consécutives
@api_router.get("/alerts/consecutive-absences")
async def calculate_consecutive_absences(
    threshold: int = 3,
    current_user: User = Depends(require_admin_or_encadrement)
):
    """Calculer les absences consécutives pour tous les cadets"""
    
    # Récupérer tous les cadets
    cadets = await db.users.find({"role": {"$in": ["cadet", "cadet_responsible"]}, "actif": True}).to_list(1000)
    
    consecutive_absences_list = []
    
    for cadet in cadets:
        # Récupérer les présences du cadet triées par date décroissante
        presences = await db.presences.find(
            {"cadet_id": cadet["id"]}
        ).sort("date", -1).to_list(1000)
        
        consecutive_count = 0
        last_absence_date = None
        
        for presence in presences:
            presence_date = datetime.fromisoformat(presence["date"]).date()
            
            if presence["status"] == "absent":
                consecutive_count += 1
                if last_absence_date is None:
                    last_absence_date = presence_date
            else:
                # Dès qu'on trouve une présence, on arrête le comptage
                break
        
        if consecutive_count >= threshold:
            consecutive_absences_list.append(ConsecutiveAbsenceCalculation(
                cadet_id=cadet["id"],
                consecutive_absences=consecutive_count,
                last_absence_date=last_absence_date
            ))
    
    return consecutive_absences_list

@api_router.get("/alerts", response_model=List[AlertResponse])
async def get_alerts(
    current_user: User = Depends(require_admin_or_encadrement)
):
    """Récupérer toutes les alertes actives"""
    
    # Récupérer les alertes depuis la base de données
    alerts = await db.alerts.find().sort("created_at", -1).to_list(1000)
    
    enriched_alerts = []
    for alert in alerts:
        # Récupérer les informations du cadet
        cadet = await db.users.find_one({"id": alert["cadet_id"]})
        if not cadet:
            continue
        
        enriched_alert = AlertResponse(
            id=alert["id"],
            cadet_id=alert["cadet_id"],
            cadet_nom=cadet["nom"],
            cadet_prenom=cadet["prenom"],
            consecutive_absences=alert["consecutive_absences"],
            last_absence_date=datetime.fromisoformat(alert["last_absence_date"]).date() if alert.get("last_absence_date") else None,
            status=AlertStatus(alert["status"]),
            contacted_by=alert.get("contacted_by"),
            contacted_at=datetime.fromisoformat(alert["contacted_at"]) if alert.get("contacted_at") else None,
            contact_comment=alert.get("contact_comment"),
            resolved_by=alert.get("resolved_by"),
            resolved_at=datetime.fromisoformat(alert["resolved_at"]) if alert.get("resolved_at") else None,
            created_at=datetime.fromisoformat(alert["created_at"])
        )
        enriched_alerts.append(enriched_alert)
    
    return enriched_alerts

@api_router.post("/alerts/generate")
async def generate_alerts(
    threshold: int = 3,
    current_user: User = Depends(require_admin_or_encadrement)
):
    """Générer de nouvelles alertes basées sur les absences consécutives"""
    
    # Calculer les absences consécutives
    consecutive_absences = await calculate_consecutive_absences(threshold, current_user)
    
    new_alerts_count = 0
    
    for absence_calc in consecutive_absences:
        # Vérifier si une alerte active existe déjà pour ce cadet
        existing_alert = await db.alerts.find_one({
            "cadet_id": absence_calc.cadet_id,
            "status": {"$in": ["active", "contacted"]}
        })
        
        if not existing_alert:
            # Créer une nouvelle alerte
            alert_data = Alert(
                cadet_id=absence_calc.cadet_id,
                consecutive_absences=absence_calc.consecutive_absences,
                last_absence_date=absence_calc.last_absence_date,
                status=AlertStatus.ACTIVE
            )
            
            # Convertir les dates en string pour MongoDB
            alert_dict = alert_data.dict()
            if alert_dict.get("last_absence_date"):
                alert_dict["last_absence_date"] = alert_dict["last_absence_date"].isoformat()
            if alert_dict.get("created_at"):
                alert_dict["created_at"] = alert_dict["created_at"].isoformat()
            
            await db.alerts.insert_one(alert_dict)
            new_alerts_count += 1
        else:
            # Mettre à jour l'alerte existante si le nombre d'absences a augmenté
            if absence_calc.consecutive_absences > existing_alert["consecutive_absences"]:
                await db.alerts.update_one(
                    {"id": existing_alert["id"]},
                    {"$set": {
                        "consecutive_absences": absence_calc.consecutive_absences,
                        "last_absence_date": absence_calc.last_absence_date.isoformat() if absence_calc.last_absence_date else None
                    }}
                )
    
    return {"message": f"{new_alerts_count} nouvelles alertes générées"}

@api_router.put("/alerts/{alert_id}")
async def update_alert(
    alert_id: str,
    alert_update: AlertUpdate,
    current_user: User = Depends(require_admin_or_encadrement)
):
    """Mettre à jour le statut d'une alerte"""
    
    # Vérifier que l'alerte existe
    existing_alert = await db.alerts.find_one({"id": alert_id})
    if not existing_alert:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Alerte non trouvée"
        )
    
    update_data = {"status": alert_update.status.value}
    
    if alert_update.status == AlertStatus.CONTACTED:
        update_data.update({
            "contacted_by": current_user.id,
            "contacted_at": datetime.utcnow().isoformat(),
            "contact_comment": alert_update.contact_comment
        })
    elif alert_update.status == AlertStatus.RESOLVED:
        update_data.update({
            "resolved_by": current_user.id,
            "resolved_at": datetime.utcnow().isoformat()
        })
    
    await db.alerts.update_one(
        {"id": alert_id},
        {"$set": update_data}
    )
    
    return {"message": "Alerte mise à jour avec succès"}

@api_router.delete("/alerts/{alert_id}")
async def delete_alert(
    alert_id: str,
    current_user: User = Depends(require_admin_or_encadrement)
):
    """Supprimer une alerte"""
    
    result = await db.alerts.delete_one({"id": alert_id})
    
    if result.deleted_count == 0:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Alerte non trouvée"
        )
    
    return {"message": "Alerte supprimée avec succès"}

# ============================================================================
# SYSTÈME DE SYNCHRONISATION HORS LIGNE
# ============================================================================

# Modèles pour la synchronisation
class OfflinePresence(BaseModel):
    """Présence enregistrée hors ligne"""
    cadet_id: str
    date: str  # Format ISO: YYYY-MM-DD
    status: PresenceStatus
    commentaire: Optional[str] = None
    timestamp: str  # Horodatage ISO avec timezone
    temp_id: str  # ID temporaire côté client pour tracking

class OfflineInspection(BaseModel):
    """Inspection d'uniforme enregistrée hors ligne"""
    cadet_id: str
    date: str  # Format ISO: YYYY-MM-DD
    uniform_type: str  # Type de tenue (ex: "C1", "C5")
    criteria_scores: Dict[str, int]  # Critère -> Score (0-4)
    commentaire: Optional[str] = None
    timestamp: str  # Horodatage ISO avec timezone
    temp_id: str

class SyncBatchRequest(BaseModel):
    """Requête de synchronisation groupée"""
    presences: List[OfflinePresence] = []
    inspections: List[OfflineInspection] = []

class SyncResult(BaseModel):
    """Résultat de synchronisation pour un élément"""
    temp_id: str
    success: bool
    server_id: Optional[str] = None  # ID créé sur le serveur
    error: Optional[str] = None
    action: str  # "created", "updated", "merged"

class SyncBatchResponse(BaseModel):
    """Réponse de synchronisation groupée"""
    presence_results: List[SyncResult] = []
    inspection_results: List[SyncResult] = []
    total_synced: int
    total_errors: int

@api_router.post("/sync/batch", response_model=SyncBatchResponse)
async def sync_offline_data(
    sync_request: SyncBatchRequest,
    current_user: User = Depends(require_presence_permissions)
):
    """
    Synchronise les données enregistrées hors ligne
    - Fusionne intelligemment les présences (la plus récente gagne)
    - Crée automatiquement une présence si inspection d'uniforme sans présence
    """
    # Log pour debug
    logger.info(f"Sync batch reçu: {len(sync_request.presences)} présences, {len(sync_request.inspections)} inspections")
    
    presence_results = []
    inspection_results = []
    
    # ========== SYNCHRONISATION DES PRÉSENCES ==========
    for offline_presence in sync_request.presences:
        try:
            # Vérifier que le cadet existe
            cadet = await db.users.find_one({"id": offline_presence.cadet_id, "actif": True})
            if not cadet:
                presence_results.append(SyncResult(
                    temp_id=offline_presence.temp_id,
                    success=False,
                    error="Cadet non trouvé",
                    action="error"
                ))
                continue
            
            # Vérifier les permissions
            if current_user.role == UserRole.CADET_RESPONSIBLE:
                if cadet.get("section_id") != current_user.section_id:
                    presence_results.append(SyncResult(
                        temp_id=offline_presence.temp_id,
                        success=False,
                        error="Permission refusée pour ce cadet",
                        action="error"
                    ))
                    continue
            
            # Chercher présence existante pour ce cadet à cette date
            existing_presence = await db.presences.find_one({
                "cadet_id": offline_presence.cadet_id,
                "date": offline_presence.date
            })
            
            if existing_presence:
                # Fusionner intelligemment : la plus récente gagne
                existing_timestamp_raw = existing_presence.get("heure_enregistrement")
                
                try:
                    # Convertir le timestamp existant en datetime avec timezone UTC
                    if isinstance(existing_timestamp_raw, str):
                        existing_timestamp = datetime.fromisoformat(existing_timestamp_raw.replace('Z', '+00:00'))
                        if existing_timestamp.tzinfo is None:
                            existing_timestamp = existing_timestamp.replace(tzinfo=timezone.utc)
                    elif isinstance(existing_timestamp_raw, datetime):
                        # Forcer la timezone UTC même s'il semble déjà avoir un tzinfo
                        if existing_timestamp_raw.tzinfo is None:
                            existing_timestamp = existing_timestamp_raw.replace(tzinfo=timezone.utc)
                        else:
                            # Convertir en UTC si timezone différente
                            existing_timestamp = existing_timestamp_raw.astimezone(timezone.utc)
                    elif existing_timestamp_raw is None:
                        # Pas de timestamp existant, utiliser un timestamp très ancien
                        existing_timestamp = datetime.min.replace(tzinfo=timezone.utc)
                    else:
                        # Type inconnu, utiliser timestamp très ancien
                        existing_timestamp = datetime.min.replace(tzinfo=timezone.utc)
                    
                    # Convertir le timestamp offline (chaîne ISO) en datetime avec timezone UTC
                    offline_timestamp = datetime.fromisoformat(offline_presence.timestamp.replace('Z', '+00:00'))
                    if offline_timestamp.tzinfo is None:
                        offline_timestamp = offline_timestamp.replace(tzinfo=timezone.utc)
                    
                    # Log pour debug
                    logger.info(f"Comparaison: existing={existing_timestamp} (tz={existing_timestamp.tzinfo}), offline={offline_timestamp} (tz={offline_timestamp.tzinfo})")
                    
                    # Comparer les timestamps
                    if offline_timestamp > existing_timestamp:
                        # La présence hors ligne est plus récente, mettre à jour
                        await db.presences.update_one(
                            {"id": existing_presence["id"]},
                            {"$set": {
                                "status": offline_presence.status.value,
                                "commentaire": offline_presence.commentaire,
                                "enregistre_par": current_user.id,
                                "heure_enregistrement": offline_timestamp.isoformat()
                            }}
                        )
                        presence_results.append(SyncResult(
                            temp_id=offline_presence.temp_id,
                            success=True,
                            server_id=existing_presence["id"],
                            action="updated"
                        ))
                    else:
                        # La présence serveur est plus récente, garder celle-ci
                        presence_results.append(SyncResult(
                            temp_id=offline_presence.temp_id,
                            success=True,
                            server_id=existing_presence["id"],
                            action="merged"
                        ))
                except Exception as e:
                    logger.error(f"Erreur comparaison timestamps: existing={existing_timestamp} ({type(existing_timestamp)}), offline={offline_presence.timestamp}, error={str(e)}")
                    presence_results.append(SyncResult(
                        temp_id=offline_presence.temp_id,
                        success=False,
                        error=str(e),
                        action="error"
                    ))
                    continue
            else:
                # Créer nouvelle présence
                presence_id = str(uuid.uuid4())
                
                # Convertir le timestamp ISO en datetime avec timezone
                offline_timestamp = datetime.fromisoformat(offline_presence.timestamp.replace('Z', '+00:00'))
                if offline_timestamp.tzinfo is None:
                    offline_timestamp = offline_timestamp.replace(tzinfo=timezone.utc)
                
                presence_data = {
                    "id": presence_id,
                    "cadet_id": offline_presence.cadet_id,
                    "date": offline_presence.date,
                    "status": offline_presence.status.value,
                    "commentaire": offline_presence.commentaire,
                    "enregistre_par": current_user.id,
                    "heure_enregistrement": offline_timestamp.isoformat(),
                    "section_id": cadet.get("section_id"),
                    "activite": None
                }
                
                await db.presences.insert_one(presence_data)
                presence_results.append(SyncResult(
                    temp_id=offline_presence.temp_id,
                    success=True,
                    server_id=presence_id,
                    action="created"
                ))
                
        except Exception as e:
            presence_results.append(SyncResult(
                temp_id=offline_presence.temp_id,
                success=False,
                error=str(e),
                action="error"
            ))
    
    # ========== SYNCHRONISATION DES INSPECTIONS D'UNIFORME ==========
    for offline_inspection in sync_request.inspections:
        try:
            # Vérifier que le cadet existe
            cadet = await db.users.find_one({"id": offline_inspection.cadet_id, "actif": True})
            if not cadet:
                inspection_results.append(SyncResult(
                    temp_id=offline_inspection.temp_id,
                    success=False,
                    error="Cadet non trouvé",
                    action="error"
                ))
                continue
            
            # Vérifier les permissions
            if current_user.role == UserRole.CADET_RESPONSIBLE:
                if cadet.get("section_id") != current_user.section_id:
                    inspection_results.append(SyncResult(
                        temp_id=offline_inspection.temp_id,
                        success=False,
                        error="Permission refusée pour ce cadet",
                        action="error"
                    ))
                    continue
            
            # LOGIQUE SPÉCIALE : Créer automatiquement une présence si elle n'existe pas
            # (cas où cadet oublie la prise de présence et va directement à l'inspection)
            existing_presence = await db.presences.find_one({
                "cadet_id": offline_inspection.cadet_id,
                "date": offline_inspection.date
            })
            
            if not existing_presence:
                # Créer une présence automatique avec statut "present"
                presence_id = str(uuid.uuid4())
                
                # Convertir le timestamp ISO en datetime avec timezone
                inspection_timestamp = datetime.fromisoformat(offline_inspection.timestamp.replace('Z', '+00:00'))
                if inspection_timestamp.tzinfo is None:
                    inspection_timestamp = inspection_timestamp.replace(tzinfo=timezone.utc)
                
                presence_data = {
                    "id": presence_id,
                    "cadet_id": offline_inspection.cadet_id,
                    "date": offline_inspection.date,
                    "status": PresenceStatus.PRESENT.value,
                    "commentaire": "Présence automatique (inspection d'uniforme)",
                    "enregistre_par": current_user.id,
                    "heure_enregistrement": inspection_timestamp.isoformat(),
                    "section_id": cadet.get("section_id"),
                    "activite": "Inspection d'uniforme"
                }
                await db.presences.insert_one(presence_data)
            
            # Vérifier s'il existe déjà une inspection pour ce cadet à cette date
            existing_inspection = await db.uniform_inspections.find_one({
                "cadet_id": offline_inspection.cadet_id,
                "date": offline_inspection.date
            })
            
            # Convertir le timestamp de l'inspection offline
            inspection_timestamp = datetime.fromisoformat(offline_inspection.timestamp.replace('Z', '+00:00'))
            if inspection_timestamp.tzinfo is None:
                inspection_timestamp = inspection_timestamp.replace(tzinfo=timezone.utc)
            
            # Si une inspection existe déjà, comparer les timestamps
            if existing_inspection:
                existing_timestamp_raw = existing_inspection.get("inspection_time")
                
                try:
                    existing_timestamp = datetime.fromisoformat(existing_timestamp_raw.replace('Z', '+00:00'))
                    if existing_timestamp.tzinfo is None:
                        existing_timestamp = existing_timestamp.replace(tzinfo=timezone.utc)
                    
                    # Si l'inspection existante est plus récente, ignorer cette sync
                    if existing_timestamp >= inspection_timestamp:
                        inspection_results.append(SyncResult(
                            temp_id=offline_inspection.temp_id,
                            success=True,
                            server_id=existing_inspection["id"],
                            action="ignored_newer_exists"
                        ))
                        continue
                    
                    # Sinon, l'inspection offline est plus ancienne (faite en premier)
                    # On écrase l'inspection existante (priorité au premier timestamp)
                    action = "updated_by_older"
                    inspection_id = existing_inspection["id"]
                    
                except (ValueError, TypeError):
                    # En cas d'erreur de parsing, garder l'existante
                    inspection_results.append(SyncResult(
                        temp_id=offline_inspection.temp_id,
                        success=True,
                        server_id=existing_inspection["id"],
                        action="ignored_timestamp_error"
                    ))
                    continue
            else:
                # Pas d'inspection existante, créer une nouvelle
                action = "created"
                inspection_id = str(uuid.uuid4())
            
            # Calculer le score total
            total_criteria = len(offline_inspection.criteria_scores)
            if total_criteria == 0:
                total_score = 0.0
                max_score = 0
            else:
                obtained_score = sum(offline_inspection.criteria_scores.values())
                max_score = total_criteria * 4
                total_score = round((obtained_score / max_score) * 100, 2) if max_score > 0 else 0.0
            
            inspection_data = {
                "id": inspection_id,
                "cadet_id": offline_inspection.cadet_id,
                "date": offline_inspection.date,
                "uniform_type": offline_inspection.uniform_type,
                "criteria_scores": offline_inspection.criteria_scores,
                "max_score": max_score,
                "total_score": total_score,
                "commentaire": offline_inspection.commentaire,
                "inspected_by": current_user.id,
                "inspection_time": inspection_timestamp.isoformat(),
                "section_id": cadet.get("section_id"),
                "auto_marked_present": not bool(existing_presence)
            }
            
            if action == "created":
                await db.uniform_inspections.insert_one(inspection_data)
            else:  # action == "updated_by_older"
                await db.uniform_inspections.replace_one(
                    {"id": inspection_id},
                    inspection_data
                )
            
            inspection_results.append(SyncResult(
                temp_id=offline_inspection.temp_id,
                success=True,
                server_id=inspection_id,
                action=action
            ))
            
        except Exception as e:
            inspection_results.append(SyncResult(
                temp_id=offline_inspection.temp_id,
                success=False,
                error=str(e),
                action="error"
            ))
    
    # Calculer les statistiques
    total_synced = sum(1 for r in presence_results + inspection_results if r.success)
    total_errors = sum(1 for r in presence_results + inspection_results if not r.success)
    
    return SyncBatchResponse(
        presence_results=presence_results,
        inspection_results=inspection_results,
        total_synced=total_synced,
        total_errors=total_errors
    )

@api_router.get("/sync/cache-data")
async def get_cache_data(current_user: User = Depends(get_current_user)):
    """
    Retourne toutes les données nécessaires pour le mode hors ligne
    - Liste des cadets actifs
    - Sections
    - Activités récentes
    """
    # Récupérer tous les cadets actifs
    users_cursor = db.users.find({"actif": True})
    users = await users_cursor.to_list(length=None)
    
    # Filtrer selon les permissions
    # Cadets avec privilèges admin peuvent voir tout le monde (comme les chefs de section)
    if current_user.has_admin_privileges:
        pass  # Pas de filtre, voir tout le monde
    elif current_user.role == UserRole.CADET_RESPONSIBLE:
        users = [u for u in users if u.get("section_id") == current_user.section_id]
    elif current_user.role == UserRole.CADET:
        users = [u for u in users if u.get("id") == current_user.id]
    
    # Récupérer les sections
    sections_cursor = db.sections.find({})
    sections = await sections_cursor.to_list(length=None)
    
    # Récupérer les activités récentes (30 derniers jours)
    thirty_days_ago = datetime.utcnow() - timedelta(days=30)
    activities_cursor = db.activities.find({
        "created_at": {"$gte": thirty_days_ago}
    })
    activities = await activities_cursor.to_list(length=None)
    
    # Nettoyer les données pour le frontend
    for user in users:
        user.pop("hashed_password", None)
        user.pop("invitation_token", None)
        if "_id" in user:
            user["_id"] = str(user["_id"])
    
    for section in sections:
        if "_id" in section:
            section["_id"] = str(section["_id"])
    
    for activity in activities:
        if "_id" in activity:
            activity["_id"] = str(activity["_id"])
    
    return {
        "users": users,
        "sections": sections,
        "activities": activities,
        "timestamp": datetime.utcnow().isoformat()
    }

# ============================================================================
# SYSTÈME D'INSPECTION DES UNIFORMES
# ============================================================================

# Modèles pour les inspections d'uniformes
class UniformSchedule(BaseModel):
    """Planification des tenues pour les inspections"""
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    date: date  # Date de l'inspection
    uniform_type: str  # Type de tenue (ex: "C1 - Tenue de Parade", "C5 - Tenue d'Entraînement")
    set_by: str  # ID de l'utilisateur qui a programmé
    set_at: datetime = Field(default_factory=datetime.utcnow)

class UniformScheduleCreate(BaseModel):
    date: date
    uniform_type: str

class Settings(BaseModel):
    """Paramètres de l'application incluant les critères d'inspection"""
    escadronName: str = ""
    address: str = ""
    contactEmail: str = ""
    allowMotivatedAbsences: bool = True
    consecutiveAbsenceThreshold: int = 3
    inspectionCriteria: Dict[str, List[str]] = {}  # Type de tenue -> liste de critères
    autoBackup: bool = True
    # Gestion des versions APK
    currentApkVersion: str = "1.0.0"
    minimumSupportedVersion: str = "1.0.0"
    apkDownloadUrl: str = ""
    forceUpdate: bool = False
    releaseNotes: List[str] = []

class UniformInspection(BaseModel):
    """Inspection d'uniforme d'un cadet"""
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    cadet_id: str
    date: date
    uniform_type: str  # Type de tenue inspecté
    criteria_scores: Dict[str, int]  # Critère -> Score (0-4)
    max_score: int  # Score maximum possible
    total_score: float  # Score total calculé (pourcentage)
    commentaire: Optional[str] = None
    inspected_by: str  # ID de l'inspecteur
    inspection_time: datetime = Field(default_factory=datetime.utcnow)
    section_id: Optional[str] = None
    auto_marked_present: bool = False  # Flag si présence créée automatiquement

class UniformInspectionCreate(BaseModel):
    cadet_id: str
    uniform_type: str
    criteria_scores: Dict[str, int]  # Critère -> Score (0-4)
    commentaire: Optional[str] = None

class UniformInspectionResponse(BaseModel):
    id: str
    cadet_id: str
    cadet_nom: str
    cadet_prenom: str
    cadet_grade: str
    date: date
    uniform_type: str
    criteria_scores: Dict[str, int]  # Critère -> Score (0-4)
    max_score: int  # Score maximum possible
    total_score: float  # Pourcentage
    commentaire: Optional[str]
    inspected_by: str
    inspector_name: str
    inspection_time: datetime
    section_id: Optional[str]
    section_nom: Optional[str]
    auto_marked_present: bool

class InspectionStatsResponse(BaseModel):
    """Statistiques d'inspection pour un cadet"""
    cadet_id: str
    cadet_nom: str
    cadet_prenom: str
    total_inspections: int
    personal_average: float  # Moyenne personnelle en pourcentage
    section_average: float   # Moyenne de la section en pourcentage
    squadron_average: float  # Moyenne de l'escadron en pourcentage
    recent_inspections: List[UniformInspectionResponse]  # 10 dernières inspections
    best_score: float
    worst_score: float

# Fonction pour vérifier les permissions d'inspection - MOVED TO LINE 493

# Fonction pour vérifier les permissions de programmation de tenue
async def require_uniform_schedule_permissions(current_user: User = Depends(get_current_user)):
    """
    Vérifie les permissions pour programmer la tenue du jour
    Autorisé: Adjudants, Adjudant-Chef, Officiers, Encadrement
    """
    # Rôles système autorisés
    allowed_system_roles = [UserRole.CADET_ADMIN, UserRole.ENCADREMENT]
    
    # Mots-clés pour les rôles personnalisés autorisés
    allowed_custom_keywords = ["adjudant", "officier", "lieutenant", "capitaine", "commandant"]
    
    # Vérifier rôles système
    if current_user.role in [r.value for r in allowed_system_roles]:
        return current_user
    
    # Vérifier rôles personnalisés par mots-clés
    role_lower = current_user.role.lower()
    if any(keyword in role_lower for keyword in allowed_custom_keywords):
        return current_user
    
    raise HTTPException(
        status_code=status.HTTP_403_FORBIDDEN,
        detail="Accès refusé. Permissions requises: Adjudants, Officiers ou Encadrement."
    )

# Routes pour les paramètres
@api_router.get("/settings", response_model=Settings)
async def get_settings(current_user: User = Depends(require_inspection_permissions)):
    """Récupérer les paramètres de l'application - accessible aux inspecteurs"""
    settings_doc = await db.settings.find_one({"type": "app_settings"})
    
    if not settings_doc:
        # Retourner les paramètres par défaut si aucun n'existe
        return Settings()
    
    # Convertir le document MongoDB en modèle Settings
    settings_dict = {
        "escadronName": settings_doc.get("escadronName", ""),
        "address": settings_doc.get("address", ""),
        "contactEmail": settings_doc.get("contactEmail", ""),
        "allowMotivatedAbsences": settings_doc.get("allowMotivatedAbsences", True),
        "consecutiveAbsenceThreshold": settings_doc.get("consecutiveAbsenceThreshold", 3),
        "inspectionCriteria": settings_doc.get("inspectionCriteria", {}),
        "autoBackup": settings_doc.get("autoBackup", True),
        "currentApkVersion": settings_doc.get("currentApkVersion", "1.0.0"),
        "minimumSupportedVersion": settings_doc.get("minimumSupportedVersion", "1.0.0"),
        "apkDownloadUrl": settings_doc.get("apkDownloadUrl", ""),
        "forceUpdate": settings_doc.get("forceUpdate", False),
        "releaseNotes": settings_doc.get("releaseNotes", [])
    }
    
    return Settings(**settings_dict)

@api_router.post("/settings")
async def save_settings(
    settings: Settings,
    current_user: User = Depends(require_admin_or_encadrement)
):
    """Sauvegarder les paramètres de l'application"""
    settings_dict = settings.dict()
    settings_dict["type"] = "app_settings"
    settings_dict["updated_by"] = current_user.id
    settings_dict["updated_at"] = datetime.utcnow().isoformat()
    
    # Upsert (update or insert)
    await db.settings.update_one(
        {"type": "app_settings"},
        {"$set": settings_dict},
        upsert=True
    )
    
    return {"message": "Paramètres sauvegardés avec succès"}

# Routes pour la planification des tenues
@api_router.get("/uniform-schedule")
async def get_uniform_schedule(
    date_param: Optional[date] = None,
    current_user: User = Depends(get_current_user)
):
    """
    Récupérer la tenue programmée pour une date
    Si aucune date n'est fournie, retourne la tenue du jour
    """
    target_date = date_param if date_param else date.today()
    
    schedule = await db.uniform_schedules.find_one({
        "date": target_date.isoformat()
    })
    
    if not schedule:
        return {
            "date": target_date.isoformat(),
            "uniform_type": None,
            "message": "Aucune tenue programmée pour cette date"
        }
    
    return {
        "id": schedule["id"],
        "date": schedule["date"],
        "uniform_type": schedule["uniform_type"],
        "set_by": schedule["set_by"],
        "set_at": schedule["set_at"]
    }

@api_router.post("/uniform-schedule")
async def set_uniform_schedule(
    schedule_data: UniformScheduleCreate,
    current_user: User = Depends(require_uniform_schedule_permissions)
):
    """
    Programmer la tenue pour une date donnée
    Autorisé: Adjudants, Officiers, Encadrement
    """
    # Vérifier si une tenue est déjà programmée pour cette date
    existing_schedule = await db.uniform_schedules.find_one({
        "date": schedule_data.date.isoformat()
    })
    
    if existing_schedule:
        # Mettre à jour la tenue existante
        await db.uniform_schedules.update_one(
            {"id": existing_schedule["id"]},
            {"$set": {
                "uniform_type": schedule_data.uniform_type,
                "set_by": current_user.id,
                "set_at": datetime.utcnow().isoformat()
            }}
        )
        return {"message": "Tenue mise à jour avec succès", "id": existing_schedule["id"]}
    else:
        # Créer une nouvelle planification
        schedule = UniformSchedule(
            date=schedule_data.date,
            uniform_type=schedule_data.uniform_type,
            set_by=current_user.id
        )
        
        schedule_dict = schedule.dict()
        schedule_dict["date"] = schedule_dict["date"].isoformat()
        schedule_dict["set_at"] = schedule_dict["set_at"].isoformat()
        
        await db.uniform_schedules.insert_one(schedule_dict)
        return {"message": "Tenue programmée avec succès", "id": schedule.id}

@api_router.delete("/uniform-schedule/{schedule_id}")
async def delete_uniform_schedule(
    schedule_id: str,
    current_user: User = Depends(require_uniform_schedule_permissions)
):
    """Supprimer une planification de tenue"""
    result = await db.uniform_schedules.delete_one({"id": schedule_id})
    
    if result.deleted_count == 0:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Planification non trouvée"
        )
    
    return {"message": "Planification supprimée avec succès"}

# Routes pour les inspections d'uniformes
@api_router.post("/uniform-inspections")
async def create_uniform_inspection(
    inspection: UniformInspectionCreate,
    inspection_date: date = None,
    current_user: User = Depends(require_inspection_permissions)
):
    """
    Créer une inspection d'uniforme
    - Calcule automatiquement le score basé sur les critères
    - Vérifie le statut de présence et marque automatiquement présent si absent/inexistant
    """
    # Utiliser la date d'aujourd'hui si non fournie
    if inspection_date is None:
        inspection_date = date.today()
    
    # Vérifier que le cadet existe
    cadet = await db.users.find_one({"id": inspection.cadet_id, "actif": True})
    if not cadet:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Cadet non trouvé"
        )
    
    # Empêcher l'auto-évaluation
    if inspection.cadet_id == current_user.id:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Vous ne pouvez pas inspecter votre propre uniforme"
        )
    
    # Vérifier les permissions selon le rôle
    user_role_lower = current_user.role.lower() if isinstance(current_user.role, str) else current_user.role.value.lower()
    
    # Déterminer si l'utilisateur est un chef de section (limité à sa section)
    is_section_leader = False
    
    # État-Major (Adjudants d'escadron) peuvent inspecter n'importe qui sauf eux-mêmes
    if 'adjudant' in user_role_lower and 'escadron' in user_role_lower:
        is_section_leader = False  # État-Major, pas de restriction de section
    
    # Chefs de section (Commandant de section, Sergent de section, Commandant de la Garde)
    elif (('commandant' in user_role_lower and 'section' in user_role_lower) or
          ('sergent' in user_role_lower and 'section' in user_role_lower) or
          ('commandant' in user_role_lower and 'garde' in user_role_lower)) and \
          current_user.section_id is not None:
        is_section_leader = True
    
    # Si c'est un chef de section, vérifier qu'il n'inspecte que sa section
    if is_section_leader:
        if cadet.get("section_id") != current_user.section_id:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="Vous ne pouvez inspecter que les cadets de votre section"
            )
    
    # Calculer le score total basé sur le barème (0-4 points par critère)
    total_criteria = len(inspection.criteria_scores)
    if total_criteria == 0:
        total_score = 0.0
        max_score = 0
    else:
        # Chaque critère est noté de 0 à 4
        obtained_score = sum(inspection.criteria_scores.values())
        max_score = total_criteria * 4  # Score maximum possible
        total_score = round((obtained_score / max_score) * 100, 2) if max_score > 0 else 0.0
    
    # Vérifier la présence du cadet pour cette date
    existing_presence = await db.presences.find_one({
        "cadet_id": inspection.cadet_id,
        "date": inspection_date.isoformat()
    })
    
    auto_marked_present = False
    
    if not existing_presence:
        # Pas de présence -> créer automatiquement une présence "present"
        presence_data = {
            "id": str(uuid.uuid4()),
            "cadet_id": inspection.cadet_id,
            "date": inspection_date.isoformat(),
            "status": "present",
            "commentaire": "Présence automatique suite à inspection uniforme",
            "enregistre_par": current_user.id,
            "heure_enregistrement": datetime.utcnow().isoformat(),
            "section_id": cadet.get("section_id"),
            "activite": f"Inspection uniforme - {inspection.uniform_type}"
        }
        await db.presences.insert_one(presence_data)
        auto_marked_present = True
    elif existing_presence.get("status") == "absent":
        # Présence marquée absente -> modifier en présent
        await db.presences.update_one(
            {"id": existing_presence["id"]},
            {"$set": {
                "status": "present",
                "commentaire": f"Modifié automatiquement suite à inspection uniforme. Ancien commentaire: {existing_presence.get('commentaire', '')}",
                "enregistre_par": current_user.id,
                "heure_enregistrement": datetime.utcnow().isoformat()
            }}
        )
        auto_marked_present = True
    
    # Créer l'inspection
    inspection_data = UniformInspection(
        cadet_id=inspection.cadet_id,
        date=inspection_date,
        uniform_type=inspection.uniform_type,
        criteria_scores=inspection.criteria_scores,
        max_score=max_score,
        total_score=total_score,
        commentaire=inspection.commentaire,
        inspected_by=current_user.id,
        section_id=cadet.get("section_id"),
        auto_marked_present=auto_marked_present
    )
    
    # Convertir en dict pour MongoDB
    inspection_dict = inspection_data.dict()
    inspection_dict["date"] = inspection_dict["date"].isoformat()
    inspection_dict["inspection_time"] = inspection_dict["inspection_time"].isoformat()
    
    await db.uniform_inspections.insert_one(inspection_dict)
    
    result = {
        "message": "Inspection enregistrée avec succès",
        "inspection_id": inspection_data.id,
        "total_score": total_score
    }
    
    if auto_marked_present:
        result["auto_marked_present"] = True
        result["presence_message"] = "Le cadet a été automatiquement marqué présent"
    
    return result

@api_router.get("/uniform-inspections", response_model=List[UniformInspectionResponse])
async def get_uniform_inspections(
    date: Optional[date] = None,
    cadet_id: Optional[str] = None,
    section_id: Optional[str] = None,
    limit: int = 100,
    current_user: User = Depends(get_current_user)
):
    """Récupérer les inspections d'uniformes avec filtres"""
    # Construire le filtre selon les permissions
    filter_dict = {}
    
    if current_user.role == UserRole.CADET:
        # Un cadet ne peut voir que ses propres inspections
        filter_dict["cadet_id"] = current_user.id
    elif current_user.role == UserRole.CADET_RESPONSIBLE:
        # Un chef de section ne peut voir que sa section
        if current_user.section_id:
            filter_dict["section_id"] = current_user.section_id
        else:
            return []
    # CADET_ADMIN et ENCADREMENT peuvent tout voir
    
    # Appliquer les filtres additionnels
    if date:
        filter_dict["date"] = date.isoformat()
    if cadet_id and current_user.role in [UserRole.CADET_ADMIN, UserRole.ENCADREMENT]:
        filter_dict["cadet_id"] = cadet_id
    if section_id and current_user.role in [UserRole.CADET_ADMIN, UserRole.ENCADREMENT]:
        filter_dict["section_id"] = section_id
    
    # Récupérer les inspections
    inspections_cursor = db.uniform_inspections.find(filter_dict).limit(limit).sort("date", -1)
    inspections = await inspections_cursor.to_list(limit)
    
    # Enrichir avec les informations des cadets, inspecteurs et sections
    enriched_inspections = []
    for inspection in inspections:
        # Récupérer les infos du cadet
        cadet = await db.users.find_one({"id": inspection["cadet_id"]})
        if not cadet:
            continue
        
        # Récupérer les infos de l'inspecteur
        inspector = await db.users.find_one({"id": inspection["inspected_by"]})
        inspector_name = f"{inspector['prenom']} {inspector['nom']}" if inspector else "Inconnu"
        
        # Récupérer les infos de la section si applicable
        section_nom = None
        if inspection.get("section_id"):
            section = await db.sections.find_one({"id": inspection["section_id"]})
            if section:
                section_nom = section["nom"]
        
        enriched_inspection = UniformInspectionResponse(
            id=inspection["id"],
            cadet_id=inspection["cadet_id"],
            cadet_nom=cadet["nom"],
            cadet_prenom=cadet["prenom"],
            cadet_grade=cadet["grade"],
            date=datetime.fromisoformat(inspection["date"]).date(),
            uniform_type=inspection["uniform_type"],
            criteria_scores=inspection["criteria_scores"],
            max_score=inspection.get("max_score", 0),
            total_score=inspection["total_score"],
            commentaire=inspection.get("commentaire"),
            inspected_by=inspection["inspected_by"],
            inspector_name=inspector_name,
            inspection_time=datetime.fromisoformat(inspection["inspection_time"]),
            section_id=inspection.get("section_id"),
            section_nom=section_nom,
            auto_marked_present=inspection.get("auto_marked_present", False)
        )
        enriched_inspections.append(enriched_inspection)
    
    return enriched_inspections

@api_router.get("/uniform-inspections/stats/me", response_model=InspectionStatsResponse)
async def get_my_inspection_stats(current_user: User = Depends(get_current_user)):
    """
    Récupérer les statistiques d'inspection personnelles d'un cadet
    et les comparer avec les moyennes de section et d'escadron
    """
    # Récupérer toutes les inspections du cadet
    my_inspections_cursor = db.uniform_inspections.find({"cadet_id": current_user.id}).sort("date", -1)
    my_inspections = await my_inspections_cursor.to_list(None)
    
    # Calculer la moyenne personnelle
    total_inspections = len(my_inspections)
    personal_average = 0.0
    best_score = 0.0
    worst_score = 100.0
    
    if total_inspections > 0:
        total_personal_score = sum(insp["total_score"] for insp in my_inspections)
        personal_average = total_personal_score / total_inspections
        best_score = max(insp["total_score"] for insp in my_inspections)
        worst_score = min(insp["total_score"] for insp in my_inspections)
    
    # Calculer la moyenne de la section
    section_average = 0.0
    if current_user.section_id:
        section_inspections_cursor = db.uniform_inspections.find({"section_id": current_user.section_id})
        section_inspections = await section_inspections_cursor.to_list(None)
        if section_inspections:
            total_section_score = sum(insp["total_score"] for insp in section_inspections)
            section_average = total_section_score / len(section_inspections)
    
    # Calculer la moyenne de l'escadron
    squadron_inspections_cursor = db.uniform_inspections.find({})
    squadron_inspections = await squadron_inspections_cursor.to_list(None)
    squadron_average = 0.0
    if squadron_inspections:
        total_squadron_score = sum(insp["total_score"] for insp in squadron_inspections)
        squadron_average = total_squadron_score / len(squadron_inspections)
    
    # Préparer les 10 dernières inspections avec enrichissement
    recent_inspections = []
    for inspection in my_inspections[:10]:
        # Récupérer les infos de l'inspecteur
        inspector = await db.users.find_one({"id": inspection["inspected_by"]})
        inspector_name = f"{inspector['prenom']} {inspector['nom']}" if inspector else "Inconnu"
        
        # Récupérer les infos de la section
        section_nom = None
        if inspection.get("section_id"):
            section = await db.sections.find_one({"id": inspection["section_id"]})
            if section:
                section_nom = section["nom"]
        
        enriched_inspection = UniformInspectionResponse(
            id=inspection["id"],
            cadet_id=inspection["cadet_id"],
            cadet_nom=current_user.nom,
            cadet_prenom=current_user.prenom,
            cadet_grade=current_user.grade,
            date=datetime.fromisoformat(inspection["date"]).date(),
            uniform_type=inspection["uniform_type"],
            criteria_scores=inspection["criteria_scores"],
            max_score=inspection.get("max_score", 0),
            total_score=inspection["total_score"],
            commentaire=inspection.get("commentaire"),
            inspected_by=inspection["inspected_by"],
            inspector_name=inspector_name,
            inspection_time=datetime.fromisoformat(inspection["inspection_time"]),
            section_id=inspection.get("section_id"),
            section_nom=section_nom,
            auto_marked_present=inspection.get("auto_marked_present", False)
        )
        recent_inspections.append(enriched_inspection)
    
    return InspectionStatsResponse(
        cadet_id=current_user.id,
        cadet_nom=current_user.nom,
        cadet_prenom=current_user.prenom,
        total_inspections=total_inspections,
        personal_average=round(personal_average, 2),
        section_average=round(section_average, 2),
        squadron_average=round(squadron_average, 2),
        recent_inspections=recent_inspections,
        best_score=round(best_score, 2),
        worst_score=round(worst_score, 2) if total_inspections > 0 else 0.0
    )

@api_router.get("/organigram/public")
async def get_public_organigram(current_user: User = Depends(get_current_user)):
    """
    Récupérer les données de l'organigrame pour tous les utilisateurs authentifiés (lecture seule)
    Retourne: users, sections, roles, subgroups
    """
    try:
        # Récupérer tous les utilisateurs actifs
        users_cursor = db.users.find({"actif": True})
        users_list = await users_cursor.to_list(1000)
        
        # Récupérer toutes les sections
        sections_cursor = db.sections.find()
        sections_list = await sections_cursor.to_list(1000)
        
        # Récupérer tous les rôles
        roles_cursor = db.roles.find()
        roles_list = await roles_cursor.to_list(1000)
        
        # Récupérer tous les sous-groupes
        all_subgroups = []
        for section in sections_list:
            subgroups_cursor = db.subgroups.find({"section_id": section["id"]})
            subgroups = await subgroups_cursor.to_list(1000)
            all_subgroups.extend(subgroups)
        
        # Supprimer le champ _id de MongoDB pour éviter les erreurs de sérialisation
        for item in users_list:
            item.pop('_id', None)
        for item in sections_list:
            item.pop('_id', None)
        for item in roles_list:
            item.pop('_id', None)
        for item in all_subgroups:
            item.pop('_id', None)
        
        return {
            "users": users_list,
            "sections": sections_list,
            "roles": roles_list,
            "subgroups": all_subgroups
        }
    except Exception as e:
        logger.error(f"Erreur lors de la récupération de l'organigrame: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Erreur lors de la récupération de l'organigrame"
        )

# Route de test
@api_router.get("/")
async def root():
    return {"message": "API Gestion Escadron Cadets - v1.0.0"}

# ============================================================================

# ============================================================================
# SYSTÈME D'IMPORT/EXPORT EXCEL
# ============================================================================

# Mapping des acronymes vers les grades du système
GRADE_MAPPING = {
    "cdt": "cadet",
    "cdt 1": "cadet_air_1re_classe",
    "cpl": "caporal",
    "cpls": "caporal_section",
    "sgt": "sergent",
    "sgts": "sergent_section",
    "adj 2": "adjudant_2e_classe",
    "adj 1": "adjudant_1re_classe",
}

# Modèles pour l'import
class ImportPreviewResponse(BaseModel):
    total_rows: int
    new_cadets: List[Dict]
    updated_cadets: List[Dict]
    errors: List[Dict]
    new_sections: List[str]

class ImportConfirmRequest(BaseModel):
    changes: List[Dict]
    create_sections: bool = True

class IndividualReportRequest(BaseModel):
    cadet_id: str
    include_presences: bool = True
    include_inspections: bool = True
    start_date: Optional[date] = None
    end_date: Optional[date] = None

# Fonction utilitaire pour générer un username unique
async def generate_unique_username(prenom: str, nom: str, db_instance) -> str:
    base_username = f"{prenom.lower().strip()}.{nom.lower().strip()}"
    base_username = base_username.replace(" ", "").replace("'", "").replace("-", "")
    
    existing = await db_instance.users.find_one({"username": base_username})
    if not existing:
        return base_username
    
    counter = 1
    while True:
        new_username = f"{base_username}{counter}"
        existing = await db_instance.users.find_one({"username": new_username})
        if not existing:
            return new_username
        counter += 1

# Fonction pour parser le fichier Excel
async def parse_excel_file(file_content: bytes) -> List[Dict]:
    try:
        df = pd.read_excel(BytesIO(file_content))
        required_cols = ['Nom', 'Prénom', 'Grade', 'Groupe']
        missing_cols = [col for col in required_cols if col not in df.columns]
        
        if missing_cols:
            raise ValueError(f"Colonnes manquantes: {', '.join(missing_cols)}")
        
        df = df[required_cols]
        df = df.dropna(subset=['Nom', 'Prénom', 'Grade', 'Groupe'])
        cadets = df.to_dict('records')
        
        return cadets
    except Exception as e:
        raise ValueError(f"Erreur lors de la lecture du fichier Excel: {str(e)}")

# Endpoint: Prévisualiser l'import Excel
@api_router.post("/import/cadets/preview")
async def preview_import_cadets(
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user)
):
    """Prévisualise l'import d'un fichier Excel de cadets"""
    # Vérifier les permissions
    if not any(role in current_user.role.lower() for role in ['cadet_admin', 'encadrement']):
        raise HTTPException(status_code=403, detail="Permissions insuffisantes")
    
    try:
        logger.info(f"Import preview - Fichier reçu: {file.filename}, type: {file.content_type}")
        content = await file.read()
        logger.info(f"Import preview - Taille du fichier: {len(content)} bytes")
        cadets_data = await parse_excel_file(content)
        
        existing_users = await db.users.find().to_list(1000)
        existing_sections = await db.sections.find().to_list(100)
        
        users_by_name = {f"{u['prenom'].lower()}.{u['nom'].lower()}": u for u in existing_users}
        sections_by_name = {s['nom'].lower(): s for s in existing_sections}
        
        new_cadets = []
        updated_cadets = []
        errors = []
        new_sections = set()
        
        for idx, cadet_data in enumerate(cadets_data, start=2):
            try:
                nom = cadet_data['Nom'].strip()
                prenom = cadet_data['Prénom'].strip()
                grade_acronym = cadet_data['Grade'].strip()
                section_name_raw = cadet_data['Groupe'].strip()
                
                # Mapper "État major" (avec espace) vers "État-Major" (avec tiret)
                # Pour correspondre au groupe virtuel déjà créé
                if section_name_raw.lower() in ['état major', 'etat major', 'état-major', 'etat-major']:
                    section_name = 'État-Major'
                else:
                    section_name = section_name_raw
                
                if grade_acronym not in GRADE_MAPPING:
                    errors.append({
                        'row': idx,
                        'nom': nom,
                        'prenom': prenom,
                        'error': f"Grade inconnu: {grade_acronym}"
                    })
                    continue
                
                grade = GRADE_MAPPING[grade_acronym]
                section_lower = section_name.lower()
                section_exists = section_lower in sections_by_name
                
                if not section_exists:
                    new_sections.add(section_name)
                
                key = f"{prenom.lower()}.{nom.lower()}"
                existing_user = users_by_name.get(key)
                
                if existing_user:
                    changes = []
                    
                    if existing_user['grade'] != grade:
                        changes.append(f"Grade: {existing_user['grade']} → {grade}")
                    
                    existing_section_name = None
                    if existing_user.get('section_id'):
                        existing_section = next((s for s in existing_sections if s['id'] == existing_user['section_id']), None)
                        if existing_section:
                            existing_section_name = existing_section['nom']
                    
                    if existing_section_name != section_name:
                        changes.append(f"Section: {existing_section_name or 'Aucune'} → {section_name}")
                    
                    if changes:
                        updated_cadets.append({
                            'row': idx,
                            'nom': nom,
                            'prenom': prenom,
                            'username': existing_user['username'],
                            'changes': changes,
                            'new_grade': grade,
                            'new_section': section_name
                        })
                else:
                    username = await generate_unique_username(prenom, nom, db)
                    new_cadets.append({
                        'row': idx,
                        'nom': nom,
                        'prenom': prenom,
                        'grade': grade,
                        'section': section_name,
                        'username': username
                    })
                    
            except Exception as e:
                errors.append({
                    'row': idx,
                    'error': f"Erreur: {str(e)}"
                })
        
        return {
            "total_rows": len(cadets_data),
            "new_cadets": new_cadets,
            "updated_cadets": updated_cadets,
            "errors": errors,
            "new_sections": list(new_sections)
        }
        
    except Exception as e:
        logger.error(f"Erreur prévisualisation import: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Erreur: {str(e)}")

# Endpoint: Confirmer l'import Excel
@api_router.post("/import/cadets/confirm")
async def confirm_import_cadets(
    request: ImportConfirmRequest,
    current_user: User = Depends(get_current_user)
):
    """Confirme et applique l'import de cadets"""
    # Vérifier les permissions
    if not any(role in current_user.role.lower() for role in ['cadet_admin', 'encadrement']):
        raise HTTPException(status_code=403, detail="Permissions insuffisantes")
    
    try:
        existing_sections = await db.sections.find().to_list(100)
        sections_by_name = {s['nom'].lower(): s for s in existing_sections}
        
        new_sections_created = []
        cadets_created = []
        cadets_updated = []
        
        if request.create_sections:
            new_section_names = set()
            for change in request.changes:
                if change.get('type') in ['new', 'update']:
                    section_name = change.get('section')
                    if section_name and section_name.lower() not in sections_by_name:
                        new_section_names.add(section_name)
            
            for section_name in new_section_names:
                section_id = str(uuid.uuid4())
                new_section = {
                    "id": section_id,
                    "nom": section_name,
                    "created_at": datetime.now().isoformat()
                }
                await db.sections.insert_one(new_section)
                sections_by_name[section_name.lower()] = new_section
                new_sections_created.append(section_name)
        
        for change in request.changes:
            change_type = change.get('type')
            
            if change_type == 'new':
                nom = change['nom']
                prenom = change['prenom']
                grade = change['grade']
                section_name = change['section']
                username = change['username']
                
                section = sections_by_name.get(section_name.lower())
                section_id = section['id'] if section else None
                
                user_id = str(uuid.uuid4())
                new_user = {
                    "id": user_id,
                    "username": username,
                    "email": None,  # Email optionnel - non utilisé pour les cadets importés
                    "nom": nom,
                    "prenom": prenom,
                    "grade": grade,
                    "role": "cadet",
                    "section_id": section_id,
                    "subgroup_id": None,
                    "photo_base64": None,
                    "actif": True,
                    "has_admin_privileges": False,
                    "hashed_password": None,
                    "must_change_password": True,
                    "created_at": datetime.now().isoformat(),
                    "invitation_token": None,
                    "invitation_expires": None,
                    "created_by": current_user.username
                }
                
                await db.users.insert_one(new_user)
                cadets_created.append(username)
                
            elif change_type == 'update':
                username = change['username']
                new_grade = change.get('new_grade')
                new_section = change.get('new_section')
                
                update_fields = {}
                
                if new_grade:
                    update_fields['grade'] = new_grade
                
                if new_section:
                    section = sections_by_name.get(new_section.lower())
                    if section:
                        update_fields['section_id'] = section['id']
                
                if update_fields:
                    await db.users.update_one(
                        {"username": username},
                        {"$set": update_fields}
                    )
                    cadets_updated.append(username)
        
        return {
            "success": True,
            "new_sections_created": new_sections_created,
            "cadets_created": len(cadets_created),
            "cadets_updated": len(cadets_updated),
            "details": {
                "created": cadets_created[:10],
                "updated": cadets_updated[:10]
            }
        }
        
    except Exception as e:
        logger.error(f"Erreur confirmation import: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Erreur: {str(e)}")

# SYSTÈME DE RAPPORTS - Endpoints pour générer des rapports
# ============================================================================

# Imports pour génération de rapports
from reportlab.lib.pagesizes import letter, A4, landscape
from reportlab.lib.units import inch
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image as ReportLabImage, PageBreak
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from reportlab.graphics.shapes import Drawing
from reportlab.graphics.charts.linecharts import HorizontalLineChart
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# Modèles pour les requêtes de rapports
class CadetsListRequest(BaseModel):
    filter_type: str = "all"  # all, section, role
    section_id: Optional[str] = None
    role: Optional[str] = None

class InspectionSheetRequest(BaseModel):
    uniform_type: str
    section_id: Optional[str] = None  # None = tous

class InspectionStatsRequest(BaseModel):
    start_date: Optional[date] = None
    end_date: Optional[date] = None
    section_id: Optional[str] = None
    include_comparisons: bool = True
    export_format: str = "pdf"  # pdf ou excel

# Configuration
LOGO_PATH = ROOT_DIR / "logo.png"

# Fonctions utilitaires pour génération PDF
def add_header_footer(canvas, doc, title: str):
    """Ajoute en-tête et pied de page à chaque page"""
    canvas.saveState()
    
    # En-tête
    if LOGO_PATH.exists():
        canvas.drawImage(str(LOGO_PATH), 0.75*inch, doc.height + 1.5*inch, width=0.75*inch, height=0.75*inch, preserveAspectRatio=True)
    
    canvas.setFont('Helvetica-Bold', 14)
    canvas.drawCentredString(doc.width/2 + doc.leftMargin, doc.height + 1.75*inch, title)
    
    canvas.setFont('Helvetica', 9)
    date_str = datetime.now().strftime("%d/%m/%Y %H:%M")
    canvas.drawCentredString(doc.width/2 + doc.leftMargin, doc.height + 1.5*inch, f"Généré le {date_str}")
    
    # Pied de page
    canvas.setFont('Helvetica', 8)
    canvas.drawCentredString(doc.width/2 + doc.leftMargin, 0.5*inch, f"Page {doc.page}")
    
    canvas.restoreState()

async def generate_cadets_list_pdf(cadets: List[dict], sections: List[dict], filter_info: str) -> BytesIO:
    """Génère un PDF avec la liste des cadets - VERSION CORRIGÉE"""
    buffer = BytesIO()
    
    # FIX #2: Augmenter les marges pour éviter que les noms soient coupés
    doc = SimpleDocTemplate(buffer, pagesize=A4, 
                            topMargin=2.5*inch, bottomMargin=0.75*inch,
                            leftMargin=1*inch, rightMargin=1*inch)  # Marges augmentées
    
    elements = []
    styles = getSampleStyleSheet()
    
    # Titre
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.HexColor('#1f2937'),
        spaceAfter=6,
        alignment=TA_CENTER
    )
    
    elements.append(Paragraph("Liste des Cadets", title_style))
    elements.append(Paragraph(f"<i>{filter_info}</i>", styles['Normal']))
    elements.append(Spacer(1, 0.3*inch))
    
    # FIX #1: Vérifier qu'il y a des cadets
    if not cadets or len(cadets) == 0:
        no_data_style = ParagraphStyle(
            'NoData',
            parent=styles['Normal'],
            fontSize=14,
            textColor=colors.HexColor('#6b7280'),
            alignment=TA_CENTER
        )
        elements.append(Paragraph("Aucun cadet trouvé avec ces critères", no_data_style))
        
        # Build PDF
        doc.build(elements, onFirstPage=lambda c, d: add_header_footer(c, d, "LISTE DES CADETS"),
                  onLaterPages=lambda c, d: add_header_footer(c, d, "LISTE DES CADETS"))
        buffer.seek(0)
        return buffer
    
    # Créer un dictionnaire des sections pour lookup rapide
    section_map = {s['id']: s['nom'] for s in sections}
    section_map['etat-major-virtual'] = '⭐ État-Major'
    
    # Grouper par section
    cadets_by_section = {}
    for cadet in cadets:
        section_id = cadet.get('section_id') or 'no_section'
        if section_id not in cadets_by_section:
            cadets_by_section[section_id] = []
        cadets_by_section[section_id].append(cadet)
    
    # Générer le tableau pour chaque section
    for section_id, section_cadets in sorted(cadets_by_section.items()):
        section_name = section_map.get(section_id, 'Sans section')
        
        # Titre de section
        section_style = ParagraphStyle(
            'SectionTitle',
            parent=styles['Heading2'],
            fontSize=14,
            textColor=colors.HexColor('#3b82f6'),
            spaceAfter=12
        )
        elements.append(Paragraph(section_name, section_style))
        
        # Tableau des cadets - Largeurs ajustées pour A4 avec marges de 1 inch
        data = [['Nom', 'Prénom', 'Grade', 'Rôle']]
        
        for cadet in sorted(section_cadets, key=lambda x: (x.get('nom') or '', x.get('prenom') or '')):
            data.append([
                cadet.get('nom', '-'),
                cadet.get('prenom', '-'),
                cadet.get('grade', '-').replace('_', ' ').title(),
                cadet.get('role', '-')
            ])
        
        # Largeurs adaptées pour A4 (210mm = 8.27 inch - 2 inch marges = 6.27 inch disponible)
        table = Table(data, colWidths=[1.5*inch, 1.5*inch, 1.6*inch, 1.6*inch])
        table.setStyle(TableStyle([
            # En-tête
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3b82f6')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            
            # Corps
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('TOPPADDING', (0, 1), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 8),
            ('LEFTPADDING', (0, 1), (-1, -1), 6),
            ('RIGHTPADDING', (0, 1), (-1, -1), 6),
            
            # Grille
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            
            # Lignes alternées
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f9fafb')])
        ]))
        
        elements.append(table)
        elements.append(Spacer(1, 0.3*inch))
    
    # Résumé
    summary_text = f"<b>Total: {len(cadets)} cadet(s)</b>"
    elements.append(Paragraph(summary_text, styles['Normal']))
    
    # Build PDF avec en-tête/pied de page
    doc.build(elements, onFirstPage=lambda c, d: add_header_footer(c, d, "LISTE DES CADETS"),
              onLaterPages=lambda c, d: add_header_footer(c, d, "LISTE DES CADETS"))
    
    buffer.seek(0)
    return buffer



async def generate_inspection_sheet_pdf(cadets: List[dict], uniform_type: str, 
                                       criteria: List[str], sections: List[dict]) -> BytesIO:
    """Génère une feuille d'inspection vierge - VERSION CORRIGÉE EN PAYSAGE"""
    buffer = BytesIO()
    
    # FIX #3: Format PAYSAGE pour plus d'espace horizontal
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4),
                            topMargin=2.5*inch, bottomMargin=0.75*inch,
                            leftMargin=0.75*inch, rightMargin=0.75*inch)
    
    elements = []
    styles = getSampleStyleSheet()
    
    # Créer un dictionnaire des sections
    section_map = {s['id']: s['nom'] for s in sections}
    section_map['etat-major-virtual'] = '⭐ État-Major'
    
    # Titre
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        textColor=colors.HexColor('#1f2937'),
        spaceAfter=6,
        alignment=TA_CENTER
    )
    
    elements.append(Paragraph(f"Feuille d'Inspection - {uniform_type}", title_style))
    elements.append(Spacer(1, 0.2*inch))
    
    # Informations
    info_style = styles['Normal']
    elements.append(Paragraph(f"<b>Date:</b> _________________________  <b>Inspecteur:</b> _________________________", info_style))
    elements.append(Spacer(1, 0.15*inch))
    
    # Légende du barème
    legend_style = ParagraphStyle(
        'Legend',
        parent=styles['Normal'],
        fontSize=9,
        textColor=colors.HexColor('#4b5563'),
        spaceAfter=10
    )
    elements.append(Paragraph("<b>Barème:</b> 0 = Très mauvais | 1 = Mauvais | 2 = Passable | 3 = Bon | 4 = Excellent", legend_style))
    elements.append(Spacer(1, 0.1*inch))
    
    # Grouper par section
    cadets_by_section = {}
    for cadet in cadets:
        section_id = cadet.get('section_id') or 'no_section'
        if section_id not in cadets_by_section:
            cadets_by_section[section_id] = []
        cadets_by_section[section_id].append(cadet)
    
    # FIX #3: Critères en abréviations pour éviter le chevauchement
    # Créer des abréviations pour les critères longs
    criteria_abbrev = []
    criteria_legend = []
    for i, crit in enumerate(criteria, 1):
        if len(crit) > 15:
            abbrev = f"C{i}"
            criteria_abbrev.append(abbrev)
            criteria_legend.append(f"{abbrev}: {crit}")
        else:
            criteria_abbrev.append(crit[:12])  # Tronquer à 12 caractères
            if len(crit) > 12:
                criteria_legend.append(f"{crit[:12]}: {crit}")
    
    # Afficher la légende des critères si nécessaire
    if criteria_legend:
        legend_text = "<b>Critères:</b> " + " | ".join(criteria_legend)
        elements.append(Paragraph(legend_text, legend_style))
        elements.append(Spacer(1, 0.1*inch))
    
    # Tableau pour chaque section
    for section_id, section_cadets in sorted(cadets_by_section.items()):
        section_name = section_map.get(section_id, 'Sans section')
        
        # Titre de section
        section_style = ParagraphStyle(
            'SectionTitle',
            parent=styles['Heading3'],
            fontSize=11,
            textColor=colors.HexColor('#3b82f6'),
            spaceAfter=6
        )
        elements.append(Paragraph(section_name, section_style))
        
        # En-tête du tableau avec abréviations
        header = ['Cadet'] + criteria_abbrev + ['Score', 'Commentaire']
        
        # Calculer les largeurs en fonction du nombre de critères
        # Paysage A4 = 11.69 inch - 1.5 inch marges = 10.19 inch disponible
        available_width = 9.5*inch
        cadet_col_width = 1.8*inch
        score_col_width = 0.5*inch
        comment_col_width = 1.5*inch
        
        remaining_width = available_width - cadet_col_width - score_col_width - comment_col_width
        criteria_col_width = remaining_width / len(criteria) if criteria else 0.5*inch
        criteria_col_width = max(criteria_col_width, 0.4*inch)  # Minimum 0.4 inch
        
        col_widths = [cadet_col_width] + [criteria_col_width] * len(criteria) + [score_col_width, comment_col_width]
        
        data = [header]
        
        # Lignes pour chaque cadet
        for cadet in sorted(section_cadets, key=lambda x: (x.get('nom') or '', x.get('prenom') or '')):
            cadet_name = f"{cadet.get('nom', '')} {cadet.get('prenom', '')}"
            row = [cadet_name] + [''] * len(criteria) + ['', '']
            data.append(row)
        
        table = Table(data, colWidths=col_widths)
        table.setStyle(TableStyle([
            # En-tête
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3b82f6')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 7),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
            ('TOPPADDING', (0, 0), (-1, 0), 6),
            
            # Corps
            ('ALIGN', (0, 1), (0, -1), 'LEFT'),
            ('ALIGN', (1, 1), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 7),
            ('TOPPADDING', (0, 1), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 8),
            
            # Grille
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            
            # Lignes alternées
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f9fafb')])
        ]))
        
        elements.append(table)
        elements.append(Spacer(1, 0.2*inch))
        
        # Page break entre sections si beaucoup de cadets
        if len(section_cadets) > 20:
            elements.append(PageBreak())
    
    # Build PDF
    doc.build(elements, onFirstPage=lambda c, d: add_header_footer(c, d, f"FEUILLE D'INSPECTION - {uniform_type.upper()}"),
              onLaterPages=lambda c, d: add_header_footer(c, d, f"FEUILLE D'INSPECTION - {uniform_type.upper()}"))
    
    buffer.seek(0)
    return buffer



async def generate_inspection_stats_pdf(inspections: List[dict], stats: dict, 
                                       period_info: str, sections: List[dict]) -> BytesIO:
    """Génère un rapport détaillé avec GRAPHIQUE d'évolution"""
    buffer = BytesIO()
    
    doc = SimpleDocTemplate(buffer, pagesize=A4,
                            topMargin=2.5*inch, bottomMargin=0.75*inch,
                            leftMargin=0.75*inch, rightMargin=0.75*inch)
    
    elements = []
    styles = getSampleStyleSheet()
    
    # Titre
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.HexColor('#1f2937'),
        spaceAfter=6,
        alignment=TA_CENTER
    )
    
    elements.append(Paragraph("Rapport d'Inspections d'Uniformes", title_style))
    elements.append(Paragraph(f"<i>{period_info}</i>", styles['Normal']))
    elements.append(Spacer(1, 0.3*inch))
    
    # Statistiques globales
    summary_style = ParagraphStyle(
        'Summary',
        parent=styles['Heading2'],
        fontSize=14,
        textColor=colors.HexColor('#3b82f6'),
        spaceAfter=12
    )
    
    elements.append(Paragraph("📊 Statistiques Globales", summary_style))
    
    summary_data = [
        ['Métrique', 'Valeur'],
        ['Total d\'inspections', str(stats.get('total_inspections', 0))],
        ['Moyenne escadron', f"{stats.get('squadron_average', 0):.1f}%"],
        ['Meilleur score', f"{stats.get('best_score', 0):.1f}%"],
        ['Score le plus bas', f"{stats.get('worst_score', 0):.1f}%"],
        ['Nombre de cadets inspectés', str(stats.get('cadets_inspected', 0))]
    ]
    
    summary_table = Table(summary_data, colWidths=[3*inch, 2*inch])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3b82f6')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
        ('ALIGN', (0, 1), (0, -1), 'LEFT'),
        ('ALIGN', (1, 1), (1, -1), 'CENTER'),
        ('FONTNAME', (0, 1), (0, -1), 'Helvetica-Bold'),
        ('FONTNAME', (1, 1), (1, -1), 'Helvetica'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
    ]))
    
    elements.append(summary_table)
    elements.append(Spacer(1, 0.3*inch))
    
    # FIX #4: GRAPHIQUE D'ÉVOLUTION
    if inspections and len(inspections) > 0:
        elements.append(Paragraph("📈 Évolution des Scores dans le Temps", summary_style))
        
        # Préparer les données pour le graphique
        # Grouper par date et calculer la moyenne
        inspections_by_date = {}
        for insp in inspections:
            date_str = insp['date']
            if date_str not in inspections_by_date:
                inspections_by_date[date_str] = []
            inspections_by_date[date_str].append(insp['total_score'])
        
        # Calculer les moyennes par date
        dates = sorted(inspections_by_date.keys())
        averages = [sum(inspections_by_date[d]) / len(inspections_by_date[d]) for d in dates]
        
        # Limiter à 15 points max pour lisibilité
        if len(dates) > 15:
            step = len(dates) // 15
            dates = dates[::step]
            averages = averages[::step]
        
        # Créer le graphique
        drawing = Drawing(400, 200)
        chart = HorizontalLineChart()
        chart.x = 50
        chart.y = 50
        chart.height = 125
        chart.width = 300
        chart.data = [averages]
        chart.lines[0].strokeColor = colors.HexColor('#3b82f6')
        chart.lines[0].strokeWidth = 2
        
        # Configurer les axes
        chart.categoryAxis.categoryNames = [d[5:] for d in dates]  # Format MM-DD
        chart.categoryAxis.labels.boxAnchor = 'n'
        chart.categoryAxis.labels.angle = 45
        chart.categoryAxis.labels.fontSize = 7
        
        chart.valueAxis.valueMin = 0
        chart.valueAxis.valueMax = 100
        chart.valueAxis.valueStep = 20
        chart.valueAxis.labels.fontSize = 8
        
        drawing.add(chart)
        elements.append(drawing)
        elements.append(Spacer(1, 0.3*inch))
    
    # Statistiques par section
    if stats.get('by_section'):
        elements.append(Paragraph("📍 Statistiques par Section", summary_style))
        
        section_data = [['Section', 'Inspections', 'Moyenne', 'Cadets']]
        
        for section_stat in stats['by_section']:
            section_data.append([
                section_stat['section_name'],
                str(section_stat['total_inspections']),
                f"{section_stat['average_score']:.1f}%",
                str(section_stat['cadets_count'])
            ])
        
        section_table = Table(section_data, colWidths=[2.5*inch, 1.5*inch, 1.5*inch, 1.5*inch])
        section_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3b82f6')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('ALIGN', (0, 1), (0, -1), 'LEFT'),
            ('ALIGN', (1, 1), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f9fafb')])
        ]))
        
        elements.append(section_table)
        elements.append(Spacer(1, 0.3*inch))
    
    # Cadets nécessitant un suivi
    if stats.get('cadets_needing_attention'):
        attention_style = ParagraphStyle(
            'Attention',
            parent=styles['Heading2'],
            fontSize=14,
            textColor=colors.HexColor('#ef4444'),
            spaceAfter=12
        )
        
        elements.append(Paragraph("⚠️ Cadets nécessitant un suivi (score < 60%)", attention_style))
        
        attention_data = [['Cadet', 'Section', 'Moyenne', 'Inspections']]
        
        for cadet in stats['cadets_needing_attention']:
            attention_data.append([
                f"{cadet['nom']} {cadet['prenom']}",
                cadet['section_name'],
                f"{cadet['average_score']:.1f}%",
                str(cadet['inspection_count'])
            ])
        
        attention_table = Table(attention_data, colWidths=[2.5*inch, 2*inch, 1.5*inch, 1.5*inch])
        attention_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#ef4444')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#fee2e2')),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('ALIGN', (0, 1), (0, -1), 'LEFT'),
            ('ALIGN', (1, 1), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ]))
        
        elements.append(attention_table)
        elements.append(Spacer(1, 0.3*inch))
    
    # Top 10 cadets
    if stats.get('top_cadets'):
        top_style = ParagraphStyle(
            'Top',
            parent=styles['Heading2'],
            fontSize=14,
            textColor=colors.HexColor('#10b981'),
            spaceAfter=12
        )
        
        elements.append(Paragraph("🏆 Top 10 Cadets", top_style))
        
        top_data = [['Rang', 'Cadet', 'Section', 'Moyenne']]
        
        for idx, cadet in enumerate(stats['top_cadets'][:10], 1):
            top_data.append([
                str(idx),
                f"{cadet['nom']} {cadet['prenom']}",
                cadet['section_name'],
                f"{cadet['average_score']:.1f}%"
            ])
        
        top_table = Table(top_data, colWidths=[0.75*inch, 2.5*inch, 2*inch, 1.5*inch])
        top_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#10b981')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('ALIGN', (0, 1), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.HexColor('#d1fae5'), colors.white])
        ]))
        
        elements.append(top_table)
    
    # NOUVEAU: Analyse des critères problématiques
    if inspections and len(inspections) > 0 and stats.get('cadets_needing_attention'):
        criteria_style = ParagraphStyle(
            'Criteria',
            parent=styles['Heading2'],
            fontSize=14,
            textColor=colors.HexColor('#f59e0b'),
            spaceAfter=12
        )
        
        elements.append(Spacer(1, 0.2*inch))
        elements.append(Paragraph("🔍 Critères à Travailler (Cadets avec scores < 60%)", criteria_style))
        
        # Analyser les critères problématiques pour chaque cadet en difficulté
        for cadet in stats['cadets_needing_attention'][:10]:  # Limiter aux 10 premiers
            cadet_inspections = [insp for insp in inspections 
                                if insp['cadet_nom'] == cadet['nom'] and insp['cadet_prenom'] == cadet['prenom']]
            
            if cadet_inspections:
                # Calculer le score moyen par critère
                criteria_scores = {}
                criteria_counts = {}
                
                for insp in cadet_inspections:
                    for criterion, score in insp.get('criteria_scores', {}).items():
                        if criterion not in criteria_scores:
                            criteria_scores[criterion] = 0
                            criteria_counts[criterion] = 0
                        criteria_scores[criterion] += score
                        criteria_counts[criterion] += 1
                
                # Calculer les moyennes et identifier les critères faibles (<50%)
                weak_criteria = []
                for criterion, total_score in criteria_scores.items():
                    avg_score = (total_score / criteria_counts[criterion]) / 4 * 100  # Score sur 4 converti en %
                    if avg_score < 50:
                        weak_criteria.append(f"{criterion} ({avg_score:.0f}%)")
                
                if weak_criteria:
                    cadet_info = f"<b>{cadet['nom']} {cadet['prenom']}</b> ({cadet['section_name']}): "
                    weak_criteria_text = ", ".join(weak_criteria)
                    
                    criteria_para_style = ParagraphStyle(
                        'CriteriaPara',
                        parent=styles['Normal'],
                        fontSize=9,
                        leftIndent=20,
                        spaceAfter=6
                    )
                    elements.append(Paragraph(cadet_info + weak_criteria_text, criteria_para_style))
    
    # NOUVEAU: Commentaires notables
    if inspections and len(inspections) > 0:
        # Filtrer les inspections avec commentaires
        inspections_with_comments = [insp for insp in inspections if insp.get('commentaire') and insp['commentaire'].strip()]
        
        if inspections_with_comments:
            comments_style = ParagraphStyle(
                'Comments',
                parent=styles['Heading2'],
                fontSize=14,
                textColor=colors.HexColor('#8b5cf6'),
                spaceAfter=12
            )
            
            elements.append(Spacer(1, 0.3*inch))
            elements.append(Paragraph("💬 Commentaires des Inspecteurs", comments_style))
            
            # Limiter aux 15 commentaires les plus récents
            recent_comments = sorted(inspections_with_comments, 
                                   key=lambda x: x.get('date', ''), 
                                   reverse=True)[:15]
            
            comments_data = [['Date', 'Cadet', 'Inspecteur', 'Commentaire']]
            
            for insp in recent_comments:
                # Limiter la longueur du commentaire pour le PDF
                comment = insp['commentaire'][:100] + '...' if len(insp['commentaire']) > 100 else insp['commentaire']
                
                comments_data.append([
                    insp['date'][5:],  # Format MM-DD
                    f"{insp['cadet_nom']} {insp['cadet_prenom']}"[:20],
                    insp['inspector_name'][:15],
                    comment
                ])
            
            comments_table = Table(comments_data, colWidths=[0.8*inch, 1.5*inch, 1.2*inch, 3.5*inch])
            comments_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#8b5cf6')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 8),
                ('BACKGROUND', (0, 1), (-1, -1), colors.white),
                ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
                ('ALIGN', (0, 1), (2, -1), 'LEFT'),
                ('ALIGN', (3, 1), (3, -1), 'LEFT'),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 7),
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ('TOPPADDING', (0, 1), (-1, -1), 6),
                ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f3f4f6')])
            ]))
            
            elements.append(comments_table)
    
    # Build PDF
    doc.build(elements, onFirstPage=lambda c, d: add_header_footer(c, d, "RAPPORT D'INSPECTIONS"),
              onLaterPages=lambda c, d: add_header_footer(c, d, "RAPPORT D'INSPECTIONS"))
    
    buffer.seek(0)
    return buffer

async def generate_inspection_stats_excel(inspections: List[dict], stats: dict, period_info: str) -> BytesIO:
    """Génère un fichier Excel avec les statistiques d'inspection"""
    buffer = BytesIO()
    workbook = openpyxl.Workbook()
    
    # Feuille 1: Vue d'ensemble
    ws_overview = workbook.active
    ws_overview.title = "Vue d'ensemble"
    
    # En-tête
    ws_overview['A1'] = "Rapport d'Inspections d'Uniformes"
    ws_overview['A1'].font = Font(bold=True, size=16, color="1F2937")
    ws_overview['A2'] = period_info
    ws_overview['A2'].font = Font(italic=True, size=11, color="6B7280")
    
    # Statistiques globales
    ws_overview['A4'] = "Statistiques Globales"
    ws_overview['A4'].font = Font(bold=True, size=14, color="3B82F6")
    
    headers_overview = ['Métrique', 'Valeur']
    for col, header in enumerate(headers_overview, 1):
        cell = ws_overview.cell(row=5, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")
    
    metrics = [
        ('Total d\'inspections', stats.get('total_inspections', 0)),
        ('Moyenne escadron', f"{stats.get('squadron_average', 0):.1f}%"),
        ('Meilleur score', f"{stats.get('best_score', 0):.1f}%"),
        ('Score le plus bas', f"{stats.get('worst_score', 0):.1f}%"),
        ('Nombre de cadets inspectés', stats.get('cadets_inspected', 0))
    ]
    
    for row, (metric, value) in enumerate(metrics, 6):
        ws_overview.cell(row=row, column=1, value=metric).font = Font(bold=True)
        ws_overview.cell(row=row, column=2, value=value)
    
    # Feuille 2: Détails des inspections
    ws_details = workbook.create_sheet("Détails")
    
    headers_details = ['Date', 'Cadet', 'Section', 'Tenue', 'Score', 'Inspecteur', 'Commentaire']
    for col, header in enumerate(headers_details, 1):
        cell = ws_details.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")
    
    for row, insp in enumerate(inspections, 2):
        ws_details.cell(row=row, column=1, value=str(insp.get('date', '-')))
        ws_details.cell(row=row, column=2, value=f"{insp.get('cadet_nom', '')} {insp.get('cadet_prenom', '')}")
        ws_details.cell(row=row, column=3, value=insp.get('section_nom', '-'))
        ws_details.cell(row=row, column=4, value=insp.get('uniform_type', '-'))
        
        score_cell = ws_details.cell(row=row, column=5, value=f"{insp.get('total_score', 0):.1f}%")
        # Coloration selon le score
        score = insp.get('total_score', 0)
        if score >= 80:
            score_cell.fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
        elif score >= 60:
            score_cell.fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
        else:
            score_cell.fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
        
        ws_details.cell(row=row, column=6, value=insp.get('inspector_name', '-'))
        ws_details.cell(row=row, column=7, value=insp.get('commentaire', ''))
    
    # Ajuster les largeurs de colonnes
    for ws in [ws_overview, ws_details]:
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    workbook.save(buffer)
    buffer.seek(0)
    return buffer

@api_router.post("/reports/cadets-list")
async def generate_cadets_list_report(
    request: CadetsListRequest,
    format: str = "pdf",  # pdf ou web
    current_user: User = Depends(require_inspection_permissions)
):
    """
    Génère une liste des cadets selon les filtres
    Accessible aux inspecteurs et supérieurs
    """
    try:
        # Construire le filtre
        filter_dict = {}
        
        if request.filter_type == "section" and request.section_id:
            filter_dict["section_id"] = request.section_id
        elif request.filter_type == "role" and request.role:
            filter_dict["role"] = request.role
        
        # Récupérer les cadets
        cadets = await db.users.find(filter_dict).to_list(1000)
        
        # Filtrer les cadets (exclure encadrement/officiers)
        filtered_cadets = [c for c in cadets if not any(
            keyword in c.get('role', '').lower() 
            for keyword in ['encadrement', 'admin', 'lieutenant', 'capitaine', 'major', 'colonel']
        )]
        
        # Récupérer les sections
        sections = await db.sections.find().to_list(100)
        
        # Créer l'info de filtre
        if request.filter_type == "section" and request.section_id:
            section = next((s for s in sections if s['id'] == request.section_id), None)
            filter_info = f"Section: {section['nom'] if section else 'Inconnue'}"
        elif request.filter_type == "role" and request.role:
            filter_info = f"Rôle: {request.role}"
        else:
            filter_info = "Tous les cadets"
        
        if format == "pdf":
            # Générer le PDF
            pdf_buffer = await generate_cadets_list_pdf(filtered_cadets, sections, filter_info)
            
            return StreamingResponse(
                pdf_buffer,
                media_type="application/pdf",
                headers={"Content-Disposition": f"attachment; filename=liste_cadets_{datetime.now().strftime('%Y%m%d')}.pdf"}
            )
        else:
            # Retourner les données pour affichage web
            return {
                "cadets": filtered_cadets,
                "sections": sections,
                "filter_info": filter_info,
                "total": len(filtered_cadets)
            }
    
    except Exception as e:
        import traceback
        logger.error(f"Erreur génération liste cadets: {str(e)}")
        logger.error(f"Traceback: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=f"Erreur lors de la génération du rapport: {str(e)}")

@api_router.post("/reports/inspection-sheet")
async def generate_inspection_sheet(
    request: InspectionSheetRequest,
    current_user: User = Depends(require_inspection_permissions)
):
    """
    Génère une feuille d'inspection vierge pour impression
    Accessible aux inspecteurs et supérieurs
    """
    try:
        # Récupérer les critères pour ce type d'uniforme
        settings_doc = await db.settings.find_one({"type": "app_settings"})
        if not settings_doc or 'inspectionCriteria' not in settings_doc:
            raise HTTPException(status_code=404, detail="Critères d'inspection non configurés")
        
        inspection_criteria = settings_doc['inspectionCriteria']
        
        if request.uniform_type not in inspection_criteria:
            raise HTTPException(status_code=404, detail=f"Critères non trouvés pour {request.uniform_type}")
        
        criteria = inspection_criteria[request.uniform_type]
        
        # Construire le filtre pour les cadets
        filter_dict = {}
        if request.section_id:
            filter_dict["section_id"] = request.section_id
        
        # Récupérer les cadets (exclure encadrement/officiers)
        all_cadets = await db.users.find(filter_dict).to_list(1000)
        filtered_cadets = [c for c in all_cadets if not any(
            keyword in c.get('role', '').lower() 
            for keyword in ['encadrement', 'admin', 'lieutenant', 'capitaine', 'major', 'colonel']
        )]
        
        # Assigner section virtuelle État-Major si nécessaire
        for cadet in filtered_cadets:
            role_lower = cadet.get('role', '').lower()
            if not cadet.get('section_id') and any(keyword in role_lower for keyword in ['adjudant d\'escadron', 'adjudant-chef d\'escadron']):
                cadet['section_id'] = 'etat-major-virtual'
        
        # Récupérer les sections
        sections = await db.sections.find().to_list(100)
        
        # Générer le PDF
        pdf_buffer = await generate_inspection_sheet_pdf(filtered_cadets, request.uniform_type, criteria, sections)
        
        return StreamingResponse(
            pdf_buffer,
            media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename=feuille_inspection_{request.uniform_type.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d')}.pdf"}
        )
    
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Erreur génération feuille inspection: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Erreur lors de la génération: {str(e)}")

@api_router.post("/reports/inspection-stats")
async def generate_inspection_stats_report(
    request: InspectionStatsRequest,
    current_user: User = Depends(require_inspection_permissions)
):
    """
    Génère un rapport détaillé des inspections avec statistiques
    Accessible aux inspecteurs et supérieurs
    """
    try:
        # Construire le filtre de dates
        filter_dict = {}
        
        start_date = request.start_date or (date.today() - timedelta(days=30))
        end_date = request.end_date or date.today()
        
        filter_dict["date"] = {
            "$gte": start_date.isoformat(),
            "$lte": end_date.isoformat()
        }
        
        if request.section_id:
            filter_dict["section_id"] = request.section_id
        
        # Récupérer les inspections
        inspections = await db.uniform_inspections.find(filter_dict).to_list(10000)
        
        if not inspections:
            raise HTTPException(status_code=404, detail="Aucune inspection trouvée pour cette période")
        
        # Enrichir les données
        users = await db.users.find().to_list(1000)
        sections = await db.sections.find().to_list(100)
        
        user_map = {u['id']: u for u in users}
        section_map = {s['id']: s['nom'] for s in sections}
        section_map['etat-major-virtual'] = '⭐ État-Major'
        
        enriched_inspections = []
        for insp in inspections:
            cadet = user_map.get(insp['cadet_id'])
            inspector = user_map.get(insp['inspected_by'])
            
            if cadet and inspector:
                enriched_inspections.append({
                    'date': insp['date'],
                    'cadet_nom': cadet['nom'],
                    'cadet_prenom': cadet['prenom'],
                    'section_id': insp.get('section_id'),
                    'section_nom': section_map.get(insp.get('section_id'), '-'),
                    'uniform_type': insp['uniform_type'],
                    'total_score': insp['total_score'],
                    'inspector_name': f"{inspector['prenom']} {inspector['nom']}",
                    'commentaire': insp.get('commentaire')
                })
        
        # Calculer les statistiques
        stats = {
            'total_inspections': len(enriched_inspections),
            'squadron_average': sum(i['total_score'] for i in enriched_inspections) / len(enriched_inspections) if enriched_inspections else 0,
            'best_score': max((i['total_score'] for i in enriched_inspections), default=0),
            'worst_score': min((i['total_score'] for i in enriched_inspections), default=0),
            'cadets_inspected': len(set(i['cadet_nom'] + i['cadet_prenom'] for i in enriched_inspections))
        }
        
        # Statistiques par section
        if request.include_comparisons:
            section_stats = {}
            for insp in enriched_inspections:
                section_id = insp['section_id']
                section_name = insp['section_nom']
                
                if section_id not in section_stats:
                    section_stats[section_id] = {
                        'section_name': section_name,
                        'total_inspections': 0,
                        'scores': [],
                        'cadets': set()
                    }
                
                section_stats[section_id]['total_inspections'] += 1
                section_stats[section_id]['scores'].append(insp['total_score'])
                section_stats[section_id]['cadets'].add(insp['cadet_nom'] + insp['cadet_prenom'])
            
            stats['by_section'] = [
                {
                    'section_name': data['section_name'],
                    'total_inspections': data['total_inspections'],
                    'average_score': sum(data['scores']) / len(data['scores']) if data['scores'] else 0,
                    'cadets_count': len(data['cadets'])
                }
                for data in section_stats.values()
            ]
            
            # Cadets nécessitant un suivi (moyenne < 60%)
            cadet_stats = {}
            for insp in enriched_inspections:
                cadet_key = (insp['cadet_nom'], insp['cadet_prenom'], insp['section_nom'])
                
                if cadet_key not in cadet_stats:
                    cadet_stats[cadet_key] = {
                        'nom': insp['cadet_nom'],
                        'prenom': insp['cadet_prenom'],
                        'section_name': insp['section_nom'],
                        'scores': [],
                        'inspection_count': 0
                    }
                
                cadet_stats[cadet_key]['scores'].append(insp['total_score'])
                cadet_stats[cadet_key]['inspection_count'] += 1
            
            # Calculer les moyennes et identifier ceux en difficulté
            cadets_needing_attention = []
            top_cadets = []
            
            for cadet_data in cadet_stats.values():
                avg_score = sum(cadet_data['scores']) / len(cadet_data['scores']) if cadet_data['scores'] else 0
                cadet_data['average_score'] = avg_score
                
                if avg_score < 60:
                    cadets_needing_attention.append(cadet_data)
                
                top_cadets.append(cadet_data)
            
            # Trier top cadets par moyenne descendante
            top_cadets.sort(key=lambda x: x['average_score'], reverse=True)
            
            stats['cadets_needing_attention'] = sorted(cadets_needing_attention, key=lambda x: x['average_score'])
            stats['top_cadets'] = top_cadets[:10]
        
        # Informations de période
        period_info = f"Période: {start_date.strftime('%d/%m/%Y')} - {end_date.strftime('%d/%m/%Y')}"
        
        # Générer le format demandé
        if request.export_format == "excel":
            excel_buffer = await generate_inspection_stats_excel(enriched_inspections, stats, period_info)
            
            return StreamingResponse(
                excel_buffer,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": f"attachment; filename=rapport_inspections_{datetime.now().strftime('%Y%m%d')}.xlsx"}
            )
        else:
            # PDF par défaut
            pdf_buffer = await generate_inspection_stats_pdf(enriched_inspections, stats, period_info, sections)
            
            return StreamingResponse(
                pdf_buffer,
                media_type="application/pdf",
                headers={"Content-Disposition": f"attachment; filename=rapport_inspections_{datetime.now().strftime('%Y%m%d')}.pdf"}
            )
    
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Erreur génération rapport inspections: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Erreur lors de la génération: {str(e)}")

# ============================================================================
# FIN SYSTÈME DE RAPPORTS

# Include router
app.include_router(api_router)

# CORS middleware
app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

# ============================================================================

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()