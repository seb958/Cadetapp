# ============================================================================
# SYSTÈME DE RAPPORTS - Endpoints et fonctions pour générer des rapports
# ============================================================================

from io import BytesIO
from typing import Optional, List
from datetime import datetime, date, timedelta
from fastapi import APIRouter, Depends, HTTPException, Response
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.units import inch
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image, PageBreak
from reportlab.lib.enums import TA_CENTER, TA_LEFT
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from pathlib import Path
import os

# Import des modèles et dépendances depuis server.py
# Ces imports doivent être ajustés selon la structure

# Modèles pour les requêtes de rapports
class CadetsListRequest(BaseModel):
    filter_type: str = "all"  # all, section, role
    section_id: Optional[str] = None
    role: Optional[str] = None

class InspectionSheetRequest(BaseModel):
    uniform_type: str
    section_id: Optional[str] = None  # None = tous

class InspectionStatsRequest(BaseModel):
    start_date: Optional[date] = None
    end_date: Optional[date] = None
    section_id: Optional[str] = None
    include_comparisons: bool = True
    export_format: str = "pdf"  # pdf ou excel

# Configuration
LOGO_PATH = Path(__file__).parent / "logo.png"

# ============================================================================
# Fonctions utilitaires pour génération PDF
# ============================================================================

def add_header_footer(canvas, doc, title: str):
    """Ajoute en-tête et pied de page à chaque page"""
    canvas.saveState()
    
    # En-tête
    if LOGO_PATH.exists():
        canvas.drawImage(str(LOGO_PATH), 0.75*inch, doc.height + 1.5*inch, width=0.75*inch, height=0.75*inch, preserveAspectRatio=True)
    
    canvas.setFont('Helvetica-Bold', 14)
    canvas.drawCentredString(doc.width/2 + doc.leftMargin, doc.height + 1.75*inch, title)
    
    canvas.setFont('Helvetica', 9)
    date_str = datetime.now().strftime("%d/%m/%Y %H:%M")
    canvas.drawCentredString(doc.width/2 + doc.leftMargin, doc.height + 1.5*inch, f"Généré le {date_str}")
    
    # Pied de page
    canvas.setFont('Helvetica', 8)
    canvas.drawCentredString(doc.width/2 + doc.leftMargin, 0.5*inch, f"Page {doc.page}")
    
    canvas.restoreState()

# ============================================================================
# Génération PDF - Liste des Cadets
# ============================================================================

async def generate_cadets_list_pdf(cadets: List[dict], sections: List[dict], filter_info: str) -> BytesIO:
    """Génère un PDF avec la liste des cadets"""
    buffer = BytesIO()
    
    doc = SimpleDocTemplate(buffer, pagesize=A4, 
                            topMargin=2.5*inch, bottomMargin=0.75*inch,
                            leftMargin=0.75*inch, rightMargin=0.75*inch)
    
    elements = []
    styles = getSampleStyleSheet()
    
    # Titre
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.HexColor('#1f2937'),
        spaceAfter=6,
        alignment=TA_CENTER
    )
    
    elements.append(Paragraph("Liste des Cadets", title_style))
    elements.append(Paragraph(f"<i>{filter_info}</i>", styles['Normal']))
    elements.append(Spacer(1, 0.3*inch))
    
    # Créer un dictionnaire des sections pour lookup rapide
    section_map = {s['id']: s['nom'] for s in sections}
    section_map['etat-major-virtual'] = '⭐ État-Major'
    
    # Grouper par section
    cadets_by_section = {}
    for cadet in cadets:
        section_id = cadet.get('section_id') or 'no_section'
        if section_id not in cadets_by_section:
            cadets_by_section[section_id] = []
        cadets_by_section[section_id].append(cadet)
    
    # Générer le tableau pour chaque section
    for section_id, section_cadets in sorted(cadets_by_section.items()):
        section_name = section_map.get(section_id, 'Sans section')
        
        # Titre de section
        section_style = ParagraphStyle(
            'SectionTitle',
            parent=styles['Heading2'],
            fontSize=14,
            textColor=colors.HexColor('#3b82f6'),
            spaceAfter=12
        )
        elements.append(Paragraph(section_name, section_style))
        
        # Tableau des cadets
        data = [['Nom', 'Prénom', 'Grade', 'Rôle']]
        
        for cadet in sorted(section_cadets, key=lambda x: (x.get('nom', ''), x.get('prenom', ''))):
            data.append([
                cadet.get('nom', '-'),
                cadet.get('prenom', '-'),
                cadet.get('grade', '-').replace('_', ' ').title(),
                cadet.get('role', '-')
            ])
        
        table = Table(data, colWidths=[2*inch, 2*inch, 2*inch, 2.5*inch])
        table.setStyle(TableStyle([
            # En-tête
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3b82f6')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 11),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            
            # Corps
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('TOPPADDING', (0, 1), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 8),
            
            # Grille
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            
            # Lignes alternées
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f9fafb')])
        ]))
        
        elements.append(table)
        elements.append(Spacer(1, 0.3*inch))
    
    # Résumé
    summary_text = f"<b>Total: {len(cadets)} cadet(s)</b>"
    elements.append(Paragraph(summary_text, styles['Normal']))
    
    # Build PDF avec en-tête/pied de page
    doc.build(elements, onFirstPage=lambda c, d: add_header_footer(c, d, "LISTE DES CADETS"),
              onLaterPages=lambda c, d: add_header_footer(c, d, "LISTE DES CADETS"))
    
    buffer.seek(0)
    return buffer

# ============================================================================
# Génération PDF - Feuille d'Inspection Vierge
# ============================================================================

async def generate_inspection_sheet_pdf(cadets: List[dict], uniform_type: str, 
                                       criteria: List[str], sections: List[dict]) -> BytesIO:
    """Génère une feuille d'inspection vierge pour impression"""
    buffer = BytesIO()
    
    doc = SimpleDocTemplate(buffer, pagesize=A4,
                            topMargin=2.5*inch, bottomMargin=0.75*inch,
                            leftMargin=0.5*inch, rightMargin=0.5*inch)
    
    elements = []
    styles = getSampleStyleSheet()
    
    # Créer un dictionnaire des sections
    section_map = {s['id']: s['nom'] for s in sections}
    section_map['etat-major-virtual'] = '⭐ État-Major'
    
    # Titre
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        textColor=colors.HexColor('#1f2937'),
        spaceAfter=6,
        alignment=TA_CENTER
    )
    
    elements.append(Paragraph(f"Feuille d'Inspection - {uniform_type}", title_style))
    elements.append(Spacer(1, 0.2*inch))
    
    # Informations
    info_style = styles['Normal']
    elements.append(Paragraph(f"<b>Date:</b> _________________________", info_style))
    elements.append(Paragraph(f"<b>Inspecteur:</b> _________________________", info_style))
    elements.append(Spacer(1, 0.2*inch))
    
    # Légende du barème
    legend_style = ParagraphStyle(
        'Legend',
        parent=styles['Normal'],
        fontSize=9,
        textColor=colors.HexColor('#4b5563'),
        spaceAfter=12
    )
    elements.append(Paragraph("<b>Barème de notation:</b> 0 = Très mauvais | 1 = Mauvais | 2 = Passable | 3 = Bon | 4 = Excellent", legend_style))
    elements.append(Spacer(1, 0.15*inch))
    
    # Grouper par section
    cadets_by_section = {}
    for cadet in cadets:
        section_id = cadet.get('section_id', 'no_section')
        if section_id not in cadets_by_section:
            cadets_by_section[section_id] = []
        cadets_by_section[section_id].append(cadet)
    
    # Tableau pour chaque section
    for section_id, section_cadets in sorted(cadets_by_section.items()):
        section_name = section_map.get(section_id, 'Sans section')
        
        # Titre de section
        section_style = ParagraphStyle(
            'SectionTitle',
            parent=styles['Heading3'],
            fontSize=12,
            textColor=colors.HexColor('#3b82f6'),
            spaceAfter=8
        )
        elements.append(Paragraph(section_name, section_style))
        
        # En-tête du tableau
        header = ['Cadet'] + criteria + ['Score', 'Commentaire']
        col_widths = [1.8*inch] + [0.4*inch] * len(criteria) + [0.5*inch, 1.5*inch]
        
        # Ajuster les largeurs si trop de critères
        if len(criteria) > 5:
            col_widths = [1.5*inch] + [0.35*inch] * len(criteria) + [0.45*inch, 1.2*inch]
        
        data = [header]
        
        # Lignes pour chaque cadet
        for cadet in sorted(section_cadets, key=lambda x: (x.get('nom', ''), x.get('prenom', ''))):
            cadet_name = f"{cadet.get('nom', '')} {cadet.get('prenom', '')}"
            row = [cadet_name] + [''] * len(criteria) + ['', '']
            data.append(row)
        
        table = Table(data, colWidths=col_widths)
        table.setStyle(TableStyle([
            # En-tête
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3b82f6')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 8),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            
            # Corps
            ('ALIGN', (0, 1), (0, -1), 'LEFT'),
            ('ALIGN', (1, 1), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('TOPPADDING', (0, 1), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 10),
            
            # Grille
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            
            # Lignes alternées
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f9fafb')])
        ]))
        
        elements.append(table)
        elements.append(Spacer(1, 0.25*inch))
        
        # Page break entre sections si nécessaire
        if len(section_cadets) > 15:
            elements.append(PageBreak())
    
    # Build PDF
    doc.build(elements, onFirstPage=lambda c, d: add_header_footer(c, d, f"FEUILLE D'INSPECTION - {uniform_type.upper()}"),
              onLaterPages=lambda c, d: add_header_footer(c, d, f"FEUILLE D'INSPECTION - {uniform_type.upper()}"))
    
    buffer.seek(0)
    return buffer

# ============================================================================
# Génération PDF - Rapport d'Inspections avec Statistiques
# ============================================================================

async def generate_inspection_stats_pdf(inspections: List[dict], stats: dict, 
                                       period_info: str, sections: List[dict]) -> BytesIO:
    """Génère un rapport détaillé des inspections avec statistiques"""
    buffer = BytesIO()
    
    doc = SimpleDocTemplate(buffer, pagesize=A4,
                            topMargin=2.5*inch, bottomMargin=0.75*inch,
                            leftMargin=0.75*inch, rightMargin=0.75*inch)
    
    elements = []
    styles = getSampleStyleSheet()
    
    # Titre
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.HexColor('#1f2937'),
        spaceAfter=6,
        alignment=TA_CENTER
    )
    
    elements.append(Paragraph("Rapport d'Inspections d'Uniformes", title_style))
    elements.append(Paragraph(f"<i>{period_info}</i>", styles['Normal']))
    elements.append(Spacer(1, 0.3*inch))
    
    # Statistiques globales
    summary_style = ParagraphStyle(
        'Summary',
        parent=styles['Heading2'],
        fontSize=14,
        textColor=colors.HexColor('#3b82f6'),
        spaceAfter=12
    )
    
    elements.append(Paragraph("📊 Statistiques Globales", summary_style))
    
    summary_data = [
        ['Métrique', 'Valeur'],
        ['Total d\'inspections', str(stats.get('total_inspections', 0))],
        ['Moyenne escadron', f"{stats.get('squadron_average', 0):.1f}%"],
        ['Meilleur score', f"{stats.get('best_score', 0):.1f}%"],
        ['Score le plus bas', f"{stats.get('worst_score', 0):.1f}%"],
        ['Nombre de cadets inspectés', str(stats.get('cadets_inspected', 0))]
    ]
    
    summary_table = Table(summary_data, colWidths=[3*inch, 2*inch])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3b82f6')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
        ('ALIGN', (0, 1), (0, -1), 'LEFT'),
        ('ALIGN', (1, 1), (1, -1), 'CENTER'),
        ('FONTNAME', (0, 1), (0, -1), 'Helvetica-Bold'),
        ('FONTNAME', (1, 1), (1, -1), 'Helvetica'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
    ]))
    
    elements.append(summary_table)
    elements.append(Spacer(1, 0.3*inch))
    
    # Statistiques par section
    if stats.get('by_section'):
        elements.append(Paragraph("📍 Statistiques par Section", summary_style))
        
        section_data = [['Section', 'Inspections', 'Moyenne', 'Cadets']]
        
        for section_stat in stats['by_section']:
            section_data.append([
                section_stat['section_name'],
                str(section_stat['total_inspections']),
                f"{section_stat['average_score']:.1f}%",
                str(section_stat['cadets_count'])
            ])
        
        section_table = Table(section_data, colWidths=[2.5*inch, 1.5*inch, 1.5*inch, 1.5*inch])
        section_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3b82f6')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('ALIGN', (0, 1), (0, -1), 'LEFT'),
            ('ALIGN', (1, 1), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f9fafb')])
        ]))
        
        elements.append(section_table)
        elements.append(Spacer(1, 0.3*inch))
    
    # Cadets nécessitant un suivi
    if stats.get('cadets_needing_attention'):
        attention_style = ParagraphStyle(
            'Attention',
            parent=styles['Heading2'],
            fontSize=14,
            textColor=colors.HexColor('#ef4444'),
            spaceAfter=12
        )
        
        elements.append(Paragraph("⚠️ Cadets nécessitant un suivi (score < 60%)", attention_style))
        
        attention_data = [['Cadet', 'Section', 'Moyenne', 'Inspections']]
        
        for cadet in stats['cadets_needing_attention']:
            attention_data.append([
                f"{cadet['nom']} {cadet['prenom']}",
                cadet['section_name'],
                f"{cadet['average_score']:.1f}%",
                str(cadet['inspection_count'])
            ])
        
        attention_table = Table(attention_data, colWidths=[2.5*inch, 2*inch, 1.5*inch, 1.5*inch])
        attention_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#ef4444')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#fee2e2')),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('ALIGN', (0, 1), (0, -1), 'LEFT'),
            ('ALIGN', (1, 1), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ]))
        
        elements.append(attention_table)
        elements.append(Spacer(1, 0.3*inch))
    
    # Top 10 cadets
    if stats.get('top_cadets'):
        top_style = ParagraphStyle(
            'Top',
            parent=styles['Heading2'],
            fontSize=14,
            textColor=colors.HexColor('#10b981'),
            spaceAfter=12
        )
        
        elements.append(Paragraph("🏆 Top 10 Cadets", top_style))
        
        top_data = [['Rang', 'Cadet', 'Section', 'Moyenne']]
        
        for idx, cadet in enumerate(stats['top_cadets'][:10], 1):
            top_data.append([
                str(idx),
                f"{cadet['nom']} {cadet['prenom']}",
                cadet['section_name'],
                f"{cadet['average_score']:.1f}%"
            ])
        
        top_table = Table(top_data, colWidths=[0.75*inch, 2.5*inch, 2*inch, 1.5*inch])
        top_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#10b981')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('ALIGN', (0, 1), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.HexColor('#d1fae5'), colors.white])
        ]))
        
        elements.append(top_table)
    
    # Build PDF
    doc.build(elements, onFirstPage=lambda c, d: add_header_footer(c, d, "RAPPORT D'INSPECTIONS"),
              onLaterPages=lambda c, d: add_header_footer(c, d, "RAPPORT D'INSPECTIONS"))
    
    buffer.seek(0)
    return buffer

# ============================================================================
# Génération Excel - Rapport d'Inspections
# ============================================================================

async def generate_inspection_stats_excel(inspections: List[dict], stats: dict, period_info: str) -> BytesIO:
    """Génère un fichier Excel avec les statistiques d'inspection"""
    buffer = BytesIO()
    workbook = openpyxl.Workbook()
    
    # Feuille 1: Vue d'ensemble
    ws_overview = workbook.active
    ws_overview.title = "Vue d'ensemble"
    
    # En-tête
    ws_overview['A1'] = "Rapport d'Inspections d'Uniformes"
    ws_overview['A1'].font = Font(bold=True, size=16, color="1F2937")
    ws_overview['A2'] = period_info
    ws_overview['A2'].font = Font(italic=True, size=11, color="6B7280")
    
    # Statistiques globales
    ws_overview['A4'] = "Statistiques Globales"
    ws_overview['A4'].font = Font(bold=True, size=14, color="3B82F6")
    
    headers_overview = ['Métrique', 'Valeur']
    for col, header in enumerate(headers_overview, 1):
        cell = ws_overview.cell(row=5, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")
    
    metrics = [
        ('Total d\'inspections', stats.get('total_inspections', 0)),
        ('Moyenne escadron', f"{stats.get('squadron_average', 0):.1f}%"),
        ('Meilleur score', f"{stats.get('best_score', 0):.1f}%"),
        ('Score le plus bas', f"{stats.get('worst_score', 0):.1f}%"),
        ('Nombre de cadets inspectés', stats.get('cadets_inspected', 0))
    ]
    
    for row, (metric, value) in enumerate(metrics, 6):
        ws_overview.cell(row=row, column=1, value=metric).font = Font(bold=True)
        ws_overview.cell(row=row, column=2, value=value)
    
    # Statistiques par section
    if stats.get('by_section'):
        start_row = len(metrics) + 8
        ws_overview.cell(row=start_row, column=1, value="Statistiques par Section").font = Font(bold=True, size=14, color="3B82F6")
        
        headers_section = ['Section', 'Inspections', 'Moyenne', 'Cadets']
        for col, header in enumerate(headers_section, 1):
            cell = ws_overview.cell(row=start_row + 1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
        
        for row, section_stat in enumerate(stats['by_section'], start_row + 2):
            ws_overview.cell(row=row, column=1, value=section_stat['section_name'])
            ws_overview.cell(row=row, column=2, value=section_stat['total_inspections'])
            ws_overview.cell(row=row, column=3, value=f"{section_stat['average_score']:.1f}%")
            ws_overview.cell(row=row, column=4, value=section_stat['cadets_count'])
    
    # Feuille 2: Détails des inspections
    ws_details = workbook.create_sheet("Détails")
    
    headers_details = ['Date', 'Cadet', 'Section', 'Tenue', 'Score', 'Inspecteur', 'Commentaire']
    for col, header in enumerate(headers_details, 1):
        cell = ws_details.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")
    
    for row, insp in enumerate(inspections, 2):
        ws_details.cell(row=row, column=1, value=str(insp.get('date', '-')))
        ws_details.cell(row=row, column=2, value=f"{insp.get('cadet_nom', '')} {insp.get('cadet_prenom', '')}")
        ws_details.cell(row=row, column=3, value=insp.get('section_nom', '-'))
        ws_details.cell(row=row, column=4, value=insp.get('uniform_type', '-'))
        
        score_cell = ws_details.cell(row=row, column=5, value=f"{insp.get('total_score', 0):.1f}%")
        # Coloration selon le score
        score = insp.get('total_score', 0)
        if score >= 80:
            score_cell.fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
        elif score >= 60:
            score_cell.fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
        else:
            score_cell.fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
        
        ws_details.cell(row=row, column=6, value=insp.get('inspector_name', '-'))
        ws_details.cell(row=row, column=7, value=insp.get('commentaire', ''))
    
    # Ajuster les largeurs de colonnes
    for ws in [ws_overview, ws_details]:
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    workbook.save(buffer)
    buffer.seek(0)
    return buffer
